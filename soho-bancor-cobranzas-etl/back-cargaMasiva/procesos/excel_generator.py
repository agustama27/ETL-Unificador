"""
Generador de archivos Excel y CSV para carga masiva CRM Bancor

Crea archivos XLSX y CSV con el formato requerido:
- Hoja: "MODELO envio"
- 13 columnas específicas
- Nombre: YYYY-MM-DD_NOMBRE_ESTUDIO.xlsx / .csv
"""
from pathlib import Path
from datetime import datetime
from typing import List, Dict
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config_catalogos import COLUMNAS_SALIDA, NOMBRE_HOJA


def crear_excel_carga_masiva(
    registros: List[Dict[str, str]],
    nombre_estudio: str,
    carpeta_salida: Path,
) -> Path:
    """
    Genera archivo XLSX con formato CRM Bancor.

    Args:
        registros: Lista de registros ya mapeados y validados
        nombre_estudio: Nombre del estudio para el nombre del archivo
        carpeta_salida: Carpeta donde guardar el archivo

    Returns:
        Path del archivo generado
    """
    # Crear workbook y obtener hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = NOMBRE_HOJA

    # Definir estilos
    estilo_encabezado = Font(bold=True)
    estilo_alineacion = Alignment(horizontal='left', vertical='center')
    estilo_borde = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Escribir encabezados
    for col_idx, columna in enumerate(COLUMNAS_SALIDA, 1):
        cell = ws.cell(row=1, column=col_idx, value=columna)
        cell.font = estilo_encabezado
        cell.alignment = estilo_alineacion

    # Escribir datos
    for row_idx, registro in enumerate(registros, 2):
        for col_idx, columna in enumerate(COLUMNAS_SALIDA, 1):
            valor = registro.get(columna, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = estilo_alineacion

    # Ajustar anchos de columna
    anchos_columna = {
        'Clase de Operación': 18,
        'Estado': 10,
        'Sub- Estado': 12,
        'CUIT': 15,
        'Cuenta': 15,
        'Desc. Acuerdo Comercial': 25,
        'Acuerdo Comercial': 18,
        'Responsable': 15,
        'Descripción': 50,
        'Persona de Contacto': 20,
        'Juzgado': 15,
        'Garante': 15,
        'Notas': 30,
    }

    for col_idx, columna in enumerate(COLUMNAS_SALIDA, 1):
        letra_columna = get_column_letter(col_idx)
        ancho = anchos_columna.get(columna, 15)
        ws.column_dimensions[letra_columna].width = ancho

    # Fijar primera fila (encabezados)
    ws.freeze_panes = 'A2'

    # Generar nombre de archivo: YYYY-MM-DD_NOMBRE_ESTUDIO.xlsx
    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    nombre_estudio_limpio = nombre_estudio.upper().replace(' ', '_')
    nombre_archivo = f"{fecha_actual}_{nombre_estudio_limpio}.xlsx"
    ruta_archivo = carpeta_salida / nombre_archivo

    # Asegurar que la carpeta existe
    carpeta_salida.mkdir(parents=True, exist_ok=True)

    # Guardar archivo
    wb.save(ruta_archivo)

    return ruta_archivo


def crear_csv_carga_masiva(
    registros: List[Dict[str, str]],
    nombre_estudio: str,
    carpeta_salida: Path,
) -> Path:
    """
    Genera archivo CSV con el mismo formato que el Excel para debugging.

    Args:
        registros: Lista de registros ya mapeados y validados
        nombre_estudio: Nombre del estudio para el nombre del archivo
        carpeta_salida: Carpeta donde guardar el archivo

    Returns:
        Path del archivo CSV generado
    """
    # Generar nombre de archivo: YYYY-MM-DD_NOMBRE_ESTUDIO.csv
    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    nombre_estudio_limpio = nombre_estudio.upper().replace(' ', '_')
    nombre_archivo = f"{fecha_actual}_{nombre_estudio_limpio}.csv"
    ruta_archivo = carpeta_salida / nombre_archivo

    # Asegurar que la carpeta existe
    carpeta_salida.mkdir(parents=True, exist_ok=True)

    # Escribir CSV con punto y coma (formato europeo, compatible con Excel)
    with open(ruta_archivo, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNAS_SALIDA, delimiter=';')
        writer.writeheader()
        writer.writerows(registros)

    return ruta_archivo


def generar_resumen_validacion(
    total_procesados: int,
    total_validos: int,
    total_errores: int,
    errores_detalle: List[str],
    max_errores_mostrar: int = 10,
) -> str:
    """
    Genera reporte de validación para mostrar al usuario.

    Args:
        total_procesados: Total de registros procesados
        total_validos: Registros que pasaron validación
        total_errores: Registros con errores
        errores_detalle: Lista de mensajes de error
        max_errores_mostrar: Máximo de errores a mostrar en detalle

    Returns:
        String formateado con el resumen
    """
    lineas = [
        "",
        "=" * 50,
        "RESUMEN DE VALIDACION",
        "=" * 50,
        f"  Total procesados:    {total_procesados}",
        f"  Registros validos:   {total_validos}",
        f"  Registros con error: {total_errores}",
    ]

    if errores_detalle:
        lineas.append("")
        lineas.append("Primeros errores encontrados:")
        for i, error in enumerate(errores_detalle[:max_errores_mostrar], 1):
            lineas.append(f"  {i}. {error}")

        if len(errores_detalle) > max_errores_mostrar:
            restantes = len(errores_detalle) - max_errores_mostrar
            lineas.append(f"  ... y {restantes} errores mas")

    lineas.append("=" * 50)

    return "\n".join(lineas)

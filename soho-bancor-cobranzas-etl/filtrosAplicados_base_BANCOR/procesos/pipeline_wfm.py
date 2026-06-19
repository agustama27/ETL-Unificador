"""Pipeline WFM para app local con historial diario y salida XLSX."""

from datetime import date, datetime
import importlib.util
from pathlib import Path
import re
import shutil
import sys
import tempfile
import unicodedata
from threading import Event
from collections.abc import Callable
from typing import Any

import pandas as pd
from openpyxl import load_workbook


ENCODINGS = ["latin-1", "iso-8859-1", "cp1252", "utf-8", "utf-16"]
COLUMNAS_REQUERIDAS = {"Cliente_BT", "MontoAdeudado", "Fecha_Entrega"}
COLUMNAS_NUMERICAS = [
    "MontoAdeudado",
    "MontoVencido",
    "SaldoCapital",
    "InteresAdeudado",
    "IVAInteresAdeudado",
    "OFERTA_Importe",
    "Deuda_vencida_Clte",
    "CapitalOriginal",
    "Compensatorio",
    "Punitorios",
]

COLUMNAS_TIPO_TEL = {
    "NumeroTelefono": "fijo",
    "NumeroTrabajo": "fijo",
    "NumeroCelular": "celular",
}
PLACEHOLDER_TELEFONO = "3519999999"
LONGITUDES_VALIDAS_TEL = {
    "fijo": (10, 12),
    "celular": (12, 13),
}

ProgressCallback = Callable[[int, str], None]


class PipelineCancelledError(Exception):
    """Senial interna para cancelacion cooperativa del pipeline."""


def obtener_carpeta_base() -> Path:
    """Devuelve carpeta de trabajo según ejecución script/exe."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    carpeta_proyecto = Path(__file__).resolve().parent.parent
    carpeta_dist = carpeta_proyecto / "dist"
    carpeta_dist.mkdir(parents=True, exist_ok=True)
    return carpeta_dist


def _timestamp_actual() -> str:
    return datetime.now().strftime("%H%M%S")


def _fecha_ddmmyyyy() -> str:
    return datetime.now().strftime("%d%m%Y")


def _build_run_context(now: datetime) -> dict[str, str]:
    """Congela tokens de fecha/hora usados por toda la corrida."""
    return {
        "fecha_ddmmyyyy": now.strftime("%d%m%Y"),
        "fecha_yyyymmdd": now.strftime("%Y%m%d"),
        "fecha_carpeta": now.strftime("%d-%m-%Y"),
        "timestamp": now.strftime("%H%M%S"),
    }


def _crear_estructura_historial(base_dir: Path, fecha_carpeta: str) -> tuple[Path, Path, Path]:
    carpeta_dia = base_dir / fecha_carpeta
    carpeta_entrada = carpeta_dia / "entrada"
    carpeta_salida = carpeta_dia / "salida"
    carpeta_entrada.mkdir(parents=True, exist_ok=True)
    carpeta_salida.mkdir(parents=True, exist_ok=True)
    return carpeta_dia, carpeta_entrada, carpeta_salida


def _crear_artifact(
    name: str,
    filename: str,
    path: Path | None,
    status: str,
    error: str | None = None,
) -> dict[str, Any]:
    return {
        "name": name,
        "filename": filename,
        "path": str(path) if path else "",
        "status": status,
        "error": error,
    }


def _determinar_status_corrida(artifacts: list[dict[str, Any]]) -> str:
    if not artifacts:
        return "failed"

    xlsx = next((item for item in artifacts if item.get("name") == "xlsx"), None)
    if not xlsx or xlsx.get("status") != "generated":
        return "failed"

    if all(item.get("status") == "generated" for item in artifacts):
        return "success"

    return "partial_failure"


def _emitir_progreso(
    progress_callback: ProgressCallback | None,
    porcentaje: int,
    mensaje: str,
) -> int:
    porcentaje_sanitizado = max(0, min(100, int(porcentaje)))
    if progress_callback is not None:
        progress_callback(porcentaje_sanitizado, mensaje)
    return porcentaje_sanitizado


def _cancelar_si_corresponde(
    cancel_event: Event | None,
    logs: list[str],
    progress_callback: ProgressCallback | None,
    porcentaje: int,
) -> None:
    if cancel_event is not None and cancel_event.is_set():
        mensaje = "Cancelado por usuario"
        logs.append(mensaje)
        _emitir_progreso(progress_callback, porcentaje, mensaje)
        raise PipelineCancelledError(mensaje)


def _serie_unica_ordenada(serie: pd.Series) -> list[str]:
    valores = [
        str(valor).strip()
        for valor in serie.tolist()
        if str(valor).strip() != ""
    ]
    return sorted(set(valores))


COLUMNAS_SALIDA_ROMAN = [
    "id_cliente_bt",
    "id_cuil",
    "id_nro_documento",
    "customer_name",
    "tel_fijo",
    "tel_laboral",
    "tel_celular",
    "txt_mail",
    "id_nro_cuenta",
    "tipo_cuenta",
    "id_sucursal_cuenta",
    "tipo_campana_ref",
    "tipo_asignacion",
    "txt_gestion_descripcion",
    "tipo_estado_cuenta",
    "fecha_gestion",
    "monto_adeudado_ars",
    "monto_entrega_ars",
    "oferta_importe",
    "resumen_productos",
    "cnt_dias_mora_max",
    "fecha_limite_oferta",
    "aplica_quita",
    "monto_quita_ars",
    "fecha_limite_quita",
]

# Columnas crudas necesarias para calcular la quita (no salen al ROMAN final)
COL_TIPO_MERCADO = "Tipo_Mercado"
COL_COMPENSATORIO = "Compensatorio"
COL_PUNITORIOS = "Punitorios"

MAPA_COLUMNAS_ROMAN = {
    "Cliente_BT": "id_cliente_bt",
    "CUIL": "id_cuil",
    "NumeroDocumento": "id_nro_documento",
    "ClienteNombre": "customer_name",
    "NumeroTelefono": "tel_fijo",
    "NumeroTrabajo": "tel_laboral",
    "NumeroCelular": "tel_celular",
    "Mail": "txt_mail",
    "Nro Cuenta": "id_nro_cuenta",
    "Cuenta": "tipo_cuenta",
    "Sucursal_Cuenta": "id_sucursal_cuenta",
    "AgrupadorProducto": "tipo_agrupador_producto",
    "Campaña_REF": "tipo_campana_ref",
    "Campana_REF": "tipo_campana_ref",
    "campana_ref": "tipo_campana_ref",
    "TipoAsignacion": "tipo_asignacion",
    "Tipo_Asignacion": "tipo_asignacion",
    "GestionDescripcion": "txt_gestion_descripcion",
    "GestionDescripción": "txt_gestion_descripcion",
    "ModuloCodigo": "id_modulo_codigo",
    "NumeroOperacion": "id_nro_operacion",
    "Dias_Mora": "cnt_dias_mora",
    "MontoAdeudado": "monto_adeudado_ars",
    "MontoVencido": "monto_vencido_ars",
    "SaldoCapital": "monto_saldo_capital_ars",
    "InteresAdeudado": "monto_interes_adeudado_ars",
    "IVAInteresAdeudado": "monto_impuesto_valor_agregado_interes_adeudado_ars",
    "OFERTA_Importe": "monto_oferta_importe",
    "AnticipoMinimo": "monto_entrega_ars",
    "Estado Cuenta": "tipo_estado_cuenta",
    "Tasa_40": "tipo_tasa_40",
    "Fecha_Gestion": "fecha_gestion",
}

COL_MONTO_VENCIDO = "MontoVencido"
COL_MONTO_ADEUDADO = "MontoAdeudado"


def _parsear_decimal(valor):
    if pd.isna(valor):
        return None
    texto = str(valor).strip()
    if texto in {"", "nan", "NaN", "None", "NaT"}:
        return None
    texto = re.sub(r"[^0-9,\.\-]", "", texto)
    if texto in {"", "-", ".", ","}:
        return None
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None


def _formatear_decimal_fijo_2(valor) -> str:
    numero = _parsear_decimal(valor)
    if numero is None:
        return ""
    return f"{numero:.2f}"


def _construir_resumen_productos(grupo: pd.DataFrame) -> str:
    columnas_orden = [col for col in ["AgrupadorProducto", "NumeroOperacion"] if col in grupo.columns]
    grupo_ordenado = grupo.sort_values(columnas_orden, kind="stable") if columnas_orden else grupo
    items = []
    for _, fila in grupo_ordenado.iterrows():
        producto = corregir_codificacion_texto(fila.get("AgrupadorProducto", ""))
        producto = str(producto).strip()
        if producto in {"", "nan", "NaN", "None", "NaT"}:
            producto = "Producto sin especificar"

        deuda = _formatear_decimal_fijo_2(fila.get(COL_MONTO_ADEUDADO))
        if deuda == "":
            deuda = "0.00"

        oferta_valor = _parsear_decimal(fila.get("OFERTA_Importe"))
        oferta_txt = "NO" if oferta_valor is None or oferta_valor <= 0 else f"{oferta_valor:.2f}"
        items.append(f"{producto} DeudaVencida:{deuda} OfertaImporte:{oferta_txt}")

    return "[" + " ; ".join(items) + "]"


def _aplicar_filtros_base(df_base: pd.DataFrame) -> pd.DataFrame:
    """Aplica filtros: MontoAdeudado > 0, EstadoDescripcion != Cancelada, SaldoCapital != 0, exclude active agreements."""
    df_work = df_base.copy()

    if "EstadoDescripcion" in df_work.columns:
        df_work["EstadoDescripcion"] = df_work["EstadoDescripcion"].astype(str).str.strip()
        df_work = df_work[df_work["EstadoDescripcion"] != "Cancelada"].copy()

    if "SaldoCapital" in df_work.columns:
        mask_saldo_cero = df_work["SaldoCapital"].notna() & (df_work["SaldoCapital"] == 0)
        df_work = df_work[~mask_saldo_cero].copy()

    df_work = df_work[df_work["MontoAdeudado"].notna() & (df_work["MontoAdeudado"] > 0)].copy()

    if "Gestion_Estado" in df_work.columns and "Fecha_Gestion" in df_work.columns:
        estados_acuerdo = ["07. Promesa de Pago Pactada", "08. Gestión de Refinanciación"]
        df_work["Gestion_Estado"] = df_work["Gestion_Estado"].astype(str).str.strip()
        df_work["Gestion_Estado"] = df_work["Gestion_Estado"].replace(["nan", "NaN", "None", "NaT"], "")
        df_work["_fecha_gestion_parsed"] = pd.to_datetime(df_work["Fecha_Gestion"], dayfirst=True, errors="coerce")

        fecha_actual = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        df_work["_dias_desde_gestion"] = (fecha_actual - df_work["_fecha_gestion_parsed"]).dt.days

        mask_acuerdo_vigente = (
            df_work["Gestion_Estado"].isin(estados_acuerdo)
            & df_work["_dias_desde_gestion"].notna()
            & (df_work["_dias_desde_gestion"] <= 7)
        )
        df_work = df_work[~mask_acuerdo_vigente].copy()
        df_work = df_work.drop(columns=["_fecha_gestion_parsed", "_dias_desde_gestion"])

    return df_work


def _consolidar_por_cliente(df_base: pd.DataFrame) -> pd.DataFrame:
    """Consolida filas por Cliente_BT: SUM numeric, MAX dias_mora, comma-concat unique ops/productos."""
    if "Cliente_BT" not in df_base.columns:
        return df_base.copy()

    def consolidar_grupo(grupo: pd.DataFrame) -> pd.Series:
        resultado = grupo.iloc[0].copy()

        for col in COLUMNAS_NUMERICAS:
            if col in grupo.columns:
                resultado[col] = grupo[col].sum()

        if "Dias_Mora" in grupo.columns:
            resultado["Dias_Mora"] = pd.to_numeric(grupo["Dias_Mora"], errors="coerce").max()

        for col in ["NumeroOperacion", "AgrupadorProducto"]:
            if col in grupo.columns:
                valores = grupo[col].dropna().astype(str)
                valores = valores[valores != ""]
                if col == "AgrupadorProducto":
                    valores = valores.apply(corregir_codificacion_texto)
                resultado[col] = ",".join(valores.unique()) if len(valores) > 0 else ""

        return resultado

    df_consolidado = (
        df_base.groupby("Cliente_BT", sort=False)
        .apply(consolidar_grupo)
        .reset_index(drop=True)
    )
    return df_consolidado


def _normalizar_telefonos_roman(df_base: pd.DataFrame) -> pd.DataFrame:
    """Aplica normalizacion de telefonos: remove hyphens, replace 3519999999, add 549 to celular, 54 to fijo."""
    df_work = df_base.copy()

    numero_ficticio = "3519999999"
    for col in ["NumeroTelefono", "NumeroTrabajo", "NumeroCelular"]:
        if col in df_work.columns:
            df_work[col] = df_work[col].astype(str)
            df_work[col] = df_work[col].str.replace("-", "", regex=False)
            df_work[col] = df_work[col].replace(numero_ficticio, "")

    if "NumeroCelular" in df_work.columns:
        df_work["NumeroCelular"] = df_work["NumeroCelular"].apply(
            lambda x: normalizar_telefono(x, "celular")
        )

    for col in ["NumeroTelefono"]:
        if col in df_work.columns:
            df_work[col] = df_work[col].apply(
                lambda x: normalizar_telefono(x, "fijo")
            )

    if "NumeroTrabajo" in df_work.columns:
        df_work["NumeroTrabajo"] = df_work["NumeroTrabajo"].astype(str)
        df_work["NumeroTrabajo"] = df_work["NumeroTrabajo"].replace(["nan", "NaN", "None", "NaT"], "")
        df_work["NumeroTrabajo"] = df_work["NumeroTrabajo"].str.replace(r"\.0+$", "", regex=True)
        df_work["NumeroTrabajo"] = df_work["NumeroTrabajo"].str.replace(r"\D", "", regex=True)
        df_work["NumeroTrabajo"] = df_work["NumeroTrabajo"].replace(numero_ficticio, "")

    return df_work


def _formatear_monto_salida(valor) -> str:
    """Formatea monto para salida: sin decimales si es entero, sino 2 decimales."""
    if pd.isna(valor):
        return ""
    try:
        numero = float(valor)
        numero_redondeado = round(numero, 2)
        if numero_redondeado.is_integer():
            return str(int(numero_redondeado))
        return f"{numero_redondeado:.2f}".rstrip("0").rstrip(".")
    except (ValueError, TypeError):
        return ""


def _normalizar_tipo_campana_ref(valor: object) -> str:
    if pd.isna(valor):
        return ""

    texto = str(valor).strip()
    if texto.lower() in {"", "nan", "none", "nat"}:
        return ""

    texto_ascii = unicodedata.normalize("NFKD", texto)
    texto_ascii = "".join(ch for ch in texto_ascii if not unicodedata.combining(ch))
    return re.sub(r"[^A-Za-z0-9]", "", texto_ascii)


def _aplicar_columnas_roman(df_base: pd.DataFrame) -> pd.DataFrame:
    """Transforma al esquema ROMAN sincronizado con back-base."""
    df_work = df_base.copy()

    if "CUIL" not in df_work.columns and "id_cuil" in df_work.columns:
        df_work["CUIL"] = df_work["id_cuil"]

    if COL_MONTO_VENCIDO not in df_work.columns and "Deuda_vencida_Clte" in df_work.columns:
        df_work[COL_MONTO_VENCIDO] = df_work["Deuda_vencida_Clte"]

    if "OFERTA_Importe" not in df_work.columns:
        df_work["OFERTA_Importe"] = pd.NA

    if "Dias_Mora" in df_work.columns:
        df_work["Dias_Mora"] = pd.to_numeric(df_work["Dias_Mora"], errors="coerce")

    if COL_MONTO_VENCIDO in df_work.columns:
        df_work[COL_MONTO_VENCIDO] = pd.to_numeric(df_work[COL_MONTO_VENCIDO], errors="coerce")

    if COL_MONTO_ADEUDADO in df_work.columns:
        df_work[COL_MONTO_ADEUDADO] = pd.to_numeric(df_work[COL_MONTO_ADEUDADO], errors="coerce")

    for _col_quita in (COL_COMPENSATORIO, COL_PUNITORIOS):
        if _col_quita in df_work.columns:
            df_work[_col_quita] = pd.to_numeric(df_work[_col_quita], errors="coerce")

    if "CUIL" not in df_work.columns and "Cliente_BT" in df_work.columns:
        df_work["CUIL"] = df_work["Cliente_BT"]

    # Fuente unica de verdad para la quita: la funcion vive en back-base y aqui se
    # importa via el loader, evitando duplicar la formula financiera.
    modulo_bb = _load_back_base_generator_module()
    calcular_quita = getattr(modulo_bb, "calcular_quita", None) if modulo_bb else None
    fecha_limite_quita_cfg = getattr(modulo_bb, "FECHA_LIMITE_QUITA", None) if modulo_bb else None

    def consolidar_grupo_roman(grupo: pd.DataFrame) -> pd.Series:
        resultado = grupo.iloc[0].copy()

        if COL_MONTO_ADEUDADO in grupo.columns:
            total = grupo[COL_MONTO_ADEUDADO].sum(min_count=1)
            resultado[COL_MONTO_ADEUDADO] = total if pd.notna(total) else 0
        elif COL_MONTO_VENCIDO in grupo.columns:
            total = grupo[COL_MONTO_VENCIDO].sum(min_count=1)
            resultado[COL_MONTO_ADEUDADO] = total if pd.notna(total) else 0

        if "AnticipoMinimo" in grupo.columns:
            anticipo = grupo["AnticipoMinimo"].dropna()
            resultado["AnticipoMinimo"] = anticipo.iloc[0] if len(anticipo) > 0 else ""
        else:
            resultado["AnticipoMinimo"] = ""

        oferta_numerica = grupo["OFERTA_Importe"].apply(_parsear_decimal)
        hay_oferta = any(v is not None and v > 0 for v in oferta_numerica.tolist())
        resultado["oferta_importe"] = "si" if hay_oferta else "no"

        if "Dias_Mora" in grupo.columns:
            resultado["Dias_Mora"] = grupo["Dias_Mora"].max()

        resultado["resumen_productos"] = _construir_resumen_productos(grupo)

        # Quita de intereses (misma formula que back-base, usando el monto consolidado)
        aplica, monto_quita = "no", None
        if callable(calcular_quita):
            tipo_mercado = ""
            if COL_TIPO_MERCADO in grupo.columns:
                tipos = grupo[COL_TIPO_MERCADO].astype(str).str.strip().str.upper()
                tipos_validos = sorted(set(tipos[~tipos.isin(["", "NAN", "NONE", "NAT"])]))
                if len(tipos_validos) > 1:
                    tipo_mercado = "__MIXTO__"
                elif tipos_validos:
                    tipo_mercado = tipos_validos[0]

            comp_total = grupo[COL_COMPENSATORIO].sum(min_count=1) if COL_COMPENSATORIO in grupo.columns else None
            punit_total = grupo[COL_PUNITORIOS].sum(min_count=1) if COL_PUNITORIOS in grupo.columns else None

            aplica, monto_quita = calcular_quita(
                tipo_mercado=tipo_mercado,
                dias_mora_max=resultado.get("Dias_Mora"),
                comp_total=comp_total,
                punit_total=punit_total,
                monto_adeudado=resultado.get(COL_MONTO_ADEUDADO),
                tiene_oferta=resultado["oferta_importe"] == "si",
            )

        resultado["aplica_quita"] = aplica
        resultado["monto_quita_ars"] = _formatear_decimal_fijo_2(monto_quita) if monto_quita is not None else ""
        resultado["fecha_limite_quita"] = (fecha_limite_quita_cfg or "") if aplica == "si" else ""
        return resultado

    clave = "CUIL" if "CUIL" in df_work.columns else "Cliente_BT"
    df_consolidado = df_work.groupby(clave, sort=False).apply(consolidar_grupo_roman).reset_index(drop=True)
    df_consolidado = _normalizar_telefonos_roman(df_consolidado)
    df_consolidado = df_consolidado.rename(columns=MAPA_COLUMNAS_ROMAN)

    renombres = {
        "monto_anticipo_minimo_ars": "monto_entrega_ars",
        "cnt_dias_mora": "cnt_dias_mora_max",
    }
    presentes = [c for c in renombres if c in df_consolidado.columns]
    if presentes:
        df_consolidado = df_consolidado.rename(columns=renombres)

    if "monto_adeudado_ars" in df_consolidado.columns:
        df_consolidado["monto_adeudado_ars"] = df_consolidado["monto_adeudado_ars"].apply(_formatear_monto_salida)

    if "monto_entrega_ars" in df_consolidado.columns:
        df_consolidado["monto_entrega_ars"] = df_consolidado["monto_entrega_ars"].apply(
            lambda x: "" if pd.isna(x) else _formatear_monto_salida(x)
        )

    if "cnt_dias_mora_max" in df_consolidado.columns:
        df_consolidado["cnt_dias_mora_max"] = df_consolidado["cnt_dias_mora_max"].apply(
            lambda x: str(int(x)) if pd.notna(x) else ""
        )

    if "oferta_importe" in df_consolidado.columns:
        df_consolidado["oferta_importe"] = (
            df_consolidado["oferta_importe"].astype(str).str.strip().str.lower().replace({"": "no", "true": "si", "false": "no"})
        )

    if "fecha_gestion" in df_consolidado.columns:
        fechas = pd.to_datetime(df_consolidado["fecha_gestion"], dayfirst=True, errors="coerce")
        df_consolidado["fecha_gestion"] = fechas.dt.strftime("%Y-%m-%d").fillna("")

    for col in COLUMNAS_SALIDA_ROMAN:
        if col not in df_consolidado.columns:
            df_consolidado[col] = ""

    df_consolidado["fecha_limite_oferta"] = "2026-06-12"

    return df_consolidado[COLUMNAS_SALIDA_ROMAN]


def _generar_csv_roman(
    df_salida: pd.DataFrame,
    carpeta_salida: Path,
    fecha_yyyymmdd: str,
    force_back_base_sync: bool = False,
) -> dict[str, Any]:
    filename = f"BANCOR_ROMAN_{fecha_yyyymmdd}.csv"
    path_salida = carpeta_salida / filename
    try:
        back_base_module = _load_back_base_generator_module()
        procesar_base_completa = getattr(back_base_module, "procesar_base_completa", None) if back_base_module else None
        if force_back_base_sync and callable(procesar_base_completa):
            procesar_base_completa()
            carpeta_back_base = Path(__file__).resolve().parents[2] / "back-base" / "base-generada" / "sin-filtros"
            candidatos = sorted(carpeta_back_base.glob("BANCOR_ROMAN_*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
            if candidatos:
                shutil.copy2(candidatos[0], path_salida)
                return _crear_artifact("roman", filename, path_salida, "generated")

        df_filtrado = _aplicar_filtros_base(df_salida)
        df_filtrado, _ = filtrar_acuerdos_vigentes(df_filtrado)
        if "Gestion_Estado" in df_filtrado.columns:
            df_filtrado = df_filtrado.drop(columns=["Gestion_Estado"])

        back_base_module = _load_back_base_generator_module()
        if back_base_module is not None:
            limpiar = getattr(back_base_module, "limpiar_numero_telefono", None)
            prefijar = getattr(back_base_module, "aplicar_prefijo_telefono", None)
            if callable(limpiar) and callable(prefijar):
                if "NumeroTelefono" in df_filtrado.columns:
                    df_filtrado["NumeroTelefono"] = df_filtrado["NumeroTelefono"].apply(limpiar)
                    df_filtrado["NumeroTelefono"] = df_filtrado["NumeroTelefono"].replace("3519999999", "")
                    df_filtrado["NumeroTelefono"] = df_filtrado["NumeroTelefono"].apply(lambda x: prefijar(x, "54"))
                if "NumeroCelular" in df_filtrado.columns:
                    df_filtrado["NumeroCelular"] = df_filtrado["NumeroCelular"].apply(limpiar)
                    df_filtrado["NumeroCelular"] = df_filtrado["NumeroCelular"].replace("3519999999", "")
                    df_filtrado["NumeroCelular"] = df_filtrado["NumeroCelular"].apply(lambda x: prefijar(x, "549"))
                if "NumeroTrabajo" in df_filtrado.columns:
                    df_filtrado["NumeroTrabajo"] = df_filtrado["NumeroTrabajo"].apply(limpiar)
                    df_filtrado["NumeroTrabajo"] = df_filtrado["NumeroTrabajo"].replace("3519999999", "")

        if "CUIL" in df_filtrado.columns:
            df_filtrado = df_filtrado.sort_index(kind="stable").copy()
            df_filtrado["_orden_original"] = range(len(df_filtrado))

            def _consolidar_base(grupo: pd.DataFrame) -> pd.Series:
                resultado = grupo.iloc[0].copy()
                if "MontoAdeudado" in grupo.columns:
                    total = grupo["MontoAdeudado"].sum(min_count=1)
                    resultado["MontoAdeudado"] = total if pd.notna(total) else 0
                elif COL_MONTO_VENCIDO in grupo.columns:
                    total = grupo[COL_MONTO_VENCIDO].sum(min_count=1)
                    resultado["MontoAdeudado"] = total if pd.notna(total) else 0

                for _col_quita in (COL_COMPENSATORIO, COL_PUNITORIOS):
                    if _col_quita in grupo.columns:
                        total_quita = pd.to_numeric(grupo[_col_quita], errors="coerce").sum(min_count=1)
                        resultado[_col_quita] = total_quita if pd.notna(total_quita) else 0

                if "AnticipoMinimo" in grupo.columns:
                    anticipo = grupo["AnticipoMinimo"].dropna()
                    resultado["AnticipoMinimo"] = anticipo.iloc[0] if len(anticipo) > 0 else ""

                if "OFERTA_Importe" in grupo.columns:
                    oferta_numerica = grupo["OFERTA_Importe"].apply(_parsear_decimal)
                    hay_oferta = any(v is not None and v > 0 for v in oferta_numerica.tolist())
                    resultado["oferta_importe"] = "si" if hay_oferta else "no"

                if "Dias_Mora" in grupo.columns:
                    resultado["Dias_Mora"] = pd.to_numeric(grupo["Dias_Mora"], errors="coerce").max()

                resultado["resumen_productos"] = _construir_resumen_productos(grupo)
                return resultado

            df_filtrado = df_filtrado.groupby("CUIL", sort=False).apply(_consolidar_base).reset_index(drop=True)
            df_filtrado, _ = deduplicar_por_telefonos_back_base(df_filtrado)
            if "_orden_original" in df_filtrado.columns:
                df_filtrado = df_filtrado.sort_values(by="_orden_original", kind="stable").reset_index(drop=True)
                df_filtrado = df_filtrado.drop(columns=["_orden_original"])

        df_final = _aplicar_columnas_roman(df_filtrado)

        df_final.to_csv(
            path_salida,
            sep=";",
            encoding="utf-8",
            index=False,
            na_rep="",
        )
        return _crear_artifact("roman", filename, path_salida, "generated")
    except Exception as exc:
        return _crear_artifact("roman", filename, path_salida, "failed", str(exc))


def _obtener_columna_telefono(df_salida: pd.DataFrame, candidatas: tuple[str, ...]) -> pd.Series:
    for columna in candidatas:
        if columna in df_salida.columns:
            return df_salida[columna]
    return pd.Series([""] * len(df_salida), index=df_salida.index)


def _generar_csv_e1kia(
    df_salida: pd.DataFrame,
    carpeta_salida: Path,
    fecha_yyyymmdd: str,
    cancel_event: Event | None = None,
) -> dict[str, Any]:
    filename = f"BANCOR_E1KIA_{fecha_yyyymmdd}_sinestrategia.csv"
    path_salida = carpeta_salida / filename
    try:
        if cancel_event is not None and cancel_event.is_set():
            raise PipelineCancelledError("Cancelado por usuario")

        serie_fijo = _obtener_columna_telefono(df_salida, ("NumeroTelefono", "numero_telefono", "tel_fijo"))
        serie_celular = _obtener_columna_telefono(df_salida, ("NumeroCelular", "numero_celular", "tel_celular"))

        def _telefono_desde_roman(valor: object, tipo: str) -> str:
            if pd.isna(valor):
                return ""
            texto = str(valor).strip()
            if texto.lower() in {"", "nan", "none", "nat"}:
                return ""
            texto = re.sub(r"\.0+$", "", texto)
            texto = re.sub(r"\D", "", texto)
            return normalizar_telefono(texto, tipo)

        usa_columnas_roman = ("tel_fijo" in df_salida.columns) or ("tel_celular" in df_salida.columns)
        if usa_columnas_roman:
            telefonos_fijos = _serie_unica_ordenada(serie_fijo.apply(lambda valor: _telefono_desde_roman(valor, "fijo")))
            telefonos_celulares = _serie_unica_ordenada(serie_celular.apply(lambda valor: _telefono_desde_roman(valor, "celular")))
        else:
            telefonos_fijos = _serie_unica_ordenada(serie_fijo.apply(lambda valor: normalizar_telefono(valor, "fijo")))
            telefonos_celulares = _serie_unica_ordenada(
                serie_celular.apply(lambda valor: normalizar_telefono(valor, "celular"))
            )

        def _clave_equivalente(numero: str) -> str:
            if numero.startswith("549"):
                return "54" + numero[3:]
            return numero

        vistos: set[str] = set()
        vistos_equivalentes: set[str] = set()
        telefonos_final: list[str] = []
        for index, numero in enumerate(telefonos_fijos):
            if index % 500 == 0 and cancel_event is not None and cancel_event.is_set():
                raise PipelineCancelledError("Cancelado por usuario")
            clave = _clave_equivalente(numero)
            if numero not in vistos and clave not in vistos_equivalentes:
                telefonos_final.append(numero)
                vistos.add(numero)
                vistos_equivalentes.add(clave)

        celulares_final: list[str] = []
        for index, numero in enumerate(telefonos_celulares):
            if index % 500 == 0 and cancel_event is not None and cancel_event.is_set():
                raise PipelineCancelledError("Cancelado por usuario")
            clave = _clave_equivalente(numero)
            if numero not in vistos and clave not in vistos_equivalentes:
                celulares_final.append(numero)
                vistos.add(numero)
                vistos_equivalentes.add(clave)

        max_filas = max(len(telefonos_final), len(celulares_final), 1)
        df_e1kia = pd.DataFrame(
            {
                "tel_fijo": telefonos_final + [""] * (max_filas - len(telefonos_final)),
                "tel_celular": celulares_final + [""] * (max_filas - len(celulares_final)),
            }
        )

        df_e1kia.to_csv(
            path_salida,
            sep=";",
            decimal=",",
            encoding="utf-8",
            index=False,
            na_rep="",
        )
        return _crear_artifact("e1kia", filename, path_salida, "generated")
    except PipelineCancelledError:
        raise
    except Exception as exc:
        return _crear_artifact("e1kia", filename, path_salida, "failed", str(exc))


def _exportar_auxiliares(
    df_salida: pd.DataFrame,
    carpeta_salida: Path,
    fecha_yyyymmdd: str,
    cancel_event: Event | None = None,
    force_back_base_sync: bool = False,
) -> list[dict[str, Any]]:
    artifacts: list[dict[str, Any]] = []
    if cancel_event is not None and cancel_event.is_set():
        raise PipelineCancelledError("Cancelado por usuario")

    try:
        artifact_roman = _generar_csv_roman(df_salida, carpeta_salida, fecha_yyyymmdd, force_back_base_sync=force_back_base_sync)
    except TypeError:
        artifact_roman = _generar_csv_roman(df_salida, carpeta_salida, fecha_yyyymmdd)
    artifacts.append(artifact_roman)
    if cancel_event is not None and cancel_event.is_set():
        raise PipelineCancelledError("Cancelado por usuario")

    df_e1kia_source = df_salida
    if artifact_roman.get("status") == "generated":
        try:
            df_e1kia_source = pd.read_csv(Path(artifact_roman["path"]), sep=";", encoding="utf-8")
        except Exception:
            df_e1kia_source = df_salida

    artifacts.append(_generar_csv_e1kia(df_e1kia_source, carpeta_salida, fecha_yyyymmdd))
    return artifacts


def _validar_variables_roman_obligatorias(path_roman: Path) -> list[str]:
    """Valida variables obligatorias/opcionales del contrato ROMAN y retorna warnings."""
    warnings: list[str] = []
    try:
        df = pd.read_csv(path_roman, sep=";", dtype=str).fillna("")
    except Exception as exc:
        return [f"ADVERTENCIA ROMAN: no se pudo leer archivo para validacion ({exc})."]

    obligatorias = [
        "customer_name",
        "monto_adeudado_ars",
        "tipo_asignacion",
        "fecha_hoy",
        "fecha_limite_sistema",
        "txt_mail",
        "cnt_dias_mora",
        "tipo_campana_ref",
        "monto_oferta_importe",
        "oferta_importe",
        "resumen_productos",
        "tipo_estado_cuenta",
        "tipo_cuenta",
        "monto_entrega_ars",
    ]

    opcionales = [
        "fecha_gestion",
        "txt_gestion_descripcion",
    ]

    alias_columnas = {
        "cnt_dias_mora": "cnt_dias_mora_max",
        "monto_oferta_importe": "oferta_importe",
    }

    for columna in obligatorias:
        col_real = columna if columna in df.columns else alias_columnas.get(columna, columna)
        if col_real not in df.columns:
            warnings.append(f"ADVERTENCIA ROMAN: falta columna obligatoria '{columna}'.")
            continue

        vacios = int((df[col_real].astype(str).str.strip() == "").sum())
        if vacios > 0:
            warnings.append(
                f"ADVERTENCIA ROMAN: columna obligatoria '{columna}' tiene {vacios} filas vacias."
            )

    for columna in opcionales:
        if columna not in df.columns:
            warnings.append(f"ADVERTENCIA ROMAN: columna opcional '{columna}' no esta presente.")

    if not warnings:
        warnings.append("VALIDACION ROMAN: OK - variables obligatorias/opcionales verificadas.")

    return warnings


def _copiar_entrada_historial(path_origen: Path, carpeta_entrada: Path, timestamp: str) -> Path:
    destino = carpeta_entrada / f"{path_origen.stem}_{timestamp}{path_origen.suffix.lower()}"
    shutil.copy2(path_origen, destino)
    return destino


def corregir_codificacion_texto(texto: str) -> str:
    """Corrige mojibake frecuente en entradas CSV."""
    if pd.isna(texto) or texto == "":
        return texto if isinstance(texto, str) else ""

    texto_str = str(texto)
    if "Ã" not in texto_str:
        return texto_str

    try:
        return texto_str.encode("latin-1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return (
            texto_str.replace("Ã¡", "á")
            .replace("Ã©", "é")
            .replace("Ã­", "í")
            .replace("Ã³", "ó")
            .replace("Ãº", "ú")
            .replace("Ã±", "ñ")
            .replace("Ã¼", "ü")
            .replace("Ãœ", "Ü")
        )


def _limpiar_nombres_columnas(df_base: pd.DataFrame) -> pd.DataFrame:
    rename_map: dict[str, str] = {}
    for col in df_base.columns:
        corregida = corregir_codificacion_texto(col)
        if corregida != col:
            rename_map[col] = corregida
    if rename_map:
        return df_base.rename(columns=rename_map)
    return df_base


def leer_archivo_entrada(path_archivo: Path) -> tuple[pd.DataFrame, str]:
    """Lee XLSX/XLS/CSV con estrategia de encoding y separador robusta."""
    extension = path_archivo.suffix.lower()
    if extension in {".xlsx", ".xls"}:
        return pd.read_excel(path_archivo), "excel"

    ultimo_error = None
    for encoding in ENCODINGS:
        for separador in [";", ","]:
            try:
                df_base = pd.read_csv(path_archivo, sep=separador, encoding=encoding, low_memory=False)
                return df_base, f"csv ({encoding}, sep='{separador}')"
            except UnicodeDecodeError as exc:
                ultimo_error = exc
                continue
            except pd.errors.ParserError as exc:
                ultimo_error = exc
                continue

    try:
        df_base = pd.read_csv(
            path_archivo,
            sep=";",
            encoding="latin-1",
            low_memory=False,
            encoding_errors="replace",
        )
        return df_base, "csv (latin-1, sep=';', errors=replace)"
    except Exception as exc:
        if ultimo_error:
            raise ValueError(f"No se pudo leer el archivo CSV: {ultimo_error}") from exc
        raise


def _normalizar_columnas_numericas(df_base: pd.DataFrame) -> pd.DataFrame:
    df_base = df_base.copy()
    for columna in COLUMNAS_NUMERICAS:
        if columna in df_base.columns:
            if df_base[columna].dtype == object:
                df_base[columna] = df_base[columna].astype(str).str.replace(",", ".", regex=False)
            df_base[columna] = pd.to_numeric(df_base[columna], errors="coerce")
    return df_base


def _normalizar_columnas_texto(df_base: pd.DataFrame) -> pd.DataFrame:
    df_base = df_base.copy()

    for campo in ["NumeroDocumento", "Nro Cuenta"]:
        if campo in df_base.columns:
            df_base[campo] = df_base[campo].astype(str)
            df_base[campo] = df_base[campo].replace(["nan", "NaN", "None", "NaT"], "")
            df_base[campo] = pd.to_numeric(df_base[campo], errors="coerce")
            df_base[campo] = df_base[campo].apply(lambda x: "" if pd.isna(x) else str(int(x)))

    numero_ficticio = "3519999999"
    for campo in ["NumeroTelefono", "NumeroTrabajo", "NumeroCelular"]:
        if campo in df_base.columns:
            df_base[campo] = df_base[campo].astype(str)
            df_base[campo] = df_base[campo].replace(["nan", "NaN", "None", "NaT"], "")
            df_base[campo] = df_base[campo].str.replace("-", "", regex=False)
            df_base[campo] = df_base[campo].replace(numero_ficticio, "")

    for campo in ["Estado Cuenta", "Tasa_40", "AgrupadorProducto", "GestionDescripción"]:
        if campo in df_base.columns:
            df_base[campo] = df_base[campo].astype(str)
            df_base[campo] = df_base[campo].replace(["nan", "NaN", "None", "NaT"], "")
            df_base[campo] = df_base[campo].apply(corregir_codificacion_texto)

    return df_base


def _remover_prefijo_15(num: str) -> str:
    """Elimina el '0' de larga distancia inicial y el '15' embebido en móviles.

    El '15' sólo puede estar entre el código de área (2–4 dígitos) y el número
    local (7 dígitos). Sólo se elimina cuando la longitud total sin '0' es 12
    y al removerlo quedan exactamente 10 dígitos (área + local).
    """
    if not num:
        return ""

    if num.startswith("0"):
        num = num[1:]

    if len(num) != 12:
        return num

    for pos in (2, 3, 4):
        if num[pos:pos + 2] == "15":
            return num[:pos] + num[pos + 2:]

    return num


def normalizar_telefono(numero: Any, tipo: str) -> str:
    """Normaliza un teléfono argentino aplicando prefijo 54/549 y validando longitud.

    - tipo='fijo'    → prefijo 54,  longitud 10–12 dígitos
    - tipo='celular' → prefijo 549, longitud 12–13 dígitos

    Elimina el '0' de larga distancia y el '15' móvil embebido. Retorna ''
    para entradas vacías, placeholder y números fuera de rango válido.
    """
    if pd.isna(numero):
        return ""

    num = str(numero).strip()
    if num in {"", "nan", "NaN", "None", "NaT", PLACEHOLDER_TELEFONO}:
        return ""

    num = re.sub(r"\.0+$", "", num)
    num = re.sub(r"\D", "", num)
    if not num:
        return ""

    if num.startswith("00"):
        num = num[2:]

    min_l, max_l = LONGITUDES_VALIDAS_TEL[tipo]

    if tipo == "celular" and num.startswith("549"):
        resultado = num
    elif tipo == "fijo" and num.startswith("54") and not num.startswith("549"):
        resultado = num
    elif num.startswith("54") and tipo == "celular":
        resultado = "549" + _remover_prefijo_15(num[2:])
    else:
        parte = _remover_prefijo_15(num)
        if tipo == "fijo":
            resultado = "54" + parte
        elif parte.startswith("9"):
            resultado = "54" + parte
        else:
            resultado = "549" + parte

    if not (min_l <= len(resultado) <= max_l):
        return ""

    return resultado


def _normalizar_columnas_telefono(df_base: pd.DataFrame) -> pd.DataFrame:
    """Aplica normalizar_telefono a las 3 columnas de teléfono del DataFrame."""
    df_base = df_base.copy()
    for campo, tipo in COLUMNAS_TIPO_TEL.items():
        if campo in df_base.columns:
            if campo == "NumeroTrabajo":
                df_base[campo] = df_base[campo].astype(str)
                df_base[campo] = df_base[campo].replace(["nan", "NaN", "None", "NaT"], "")
                df_base[campo] = df_base[campo].str.replace(r"\.0+$", "", regex=True)
                df_base[campo] = df_base[campo].str.replace(r"\D", "", regex=True)
                df_base[campo] = df_base[campo].replace(PLACEHOLDER_TELEFONO, "")
            else:
                df_base[campo] = df_base[campo].apply(lambda x: normalizar_telefono(x, tipo))
    return df_base


def filtrar_fecha_entrega_por_meses(
    df_base: pd.DataFrame,
    meses_permitidos: list[int],
    incluir_fechas_vacias: bool = False,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """Aplica filtro de Fecha_Entrega según meses elegidos en UI."""
    if "Fecha_Entrega" not in df_base.columns:
        raise ValueError("Falta columna obligatoria 'Fecha_Entrega'.")

    mapa_meses = {
        "ene": "01",
        "enero": "01",
        "feb": "02",
        "febrero": "02",
        "mar": "03",
        "marzo": "03",
        "abr": "04",
        "abril": "04",
        "may": "05",
        "mayo": "05",
        "jun": "06",
        "junio": "06",
        "jul": "07",
        "julio": "07",
        "ago": "08",
        "agosto": "08",
        "sep": "09",
        "sept": "09",
        "septiembre": "09",
        "oct": "10",
        "octubre": "10",
        "nov": "11",
        "noviembre": "11",
        "dic": "12",
        "diciembre": "12",
    }

    def parsear_fecha(valor: object) -> pd.Timestamp:
        if pd.isna(valor):
            return pd.NaT
        if isinstance(valor, (pd.Timestamp, datetime, date)):
            return pd.Timestamp(valor)
        if isinstance(valor, (int, float)) and not isinstance(valor, bool):
            serial = float(valor)
            return pd.to_datetime(pd.Series([serial]), unit="D", origin="1899-12-30", errors="coerce").iloc[0]

        texto = str(valor).strip()
        if texto in {"", "nan", "NaN", "None", "NaT"}:
            return pd.NaT

        if re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}", texto):
            return pd.to_datetime(texto, errors="coerce")

        try:
            serial = float(texto.replace(",", "."))
            return pd.to_datetime(pd.Series([serial]), unit="D", origin="1899-12-30", errors="coerce").iloc[0]
        except ValueError:
            pass

        texto_normalizado = texto.lower()
        for mes_txt, mes_nro in mapa_meses.items():
            texto_normalizado = re.sub(rf"\b{mes_txt}\b", mes_nro, texto_normalizado)

        return pd.to_datetime(texto_normalizado, dayfirst=True, errors="coerce")

    fecha_parsed = df_base["Fecha_Entrega"].apply(parsear_fecha)
    mask_valida = fecha_parsed.notna()
    mask_mes = fecha_parsed.dt.month.isin(meses_permitidos)
    if incluir_fechas_vacias:
        mask_final = (mask_valida & mask_mes) | (~mask_valida)
    else:
        mask_final = mask_valida & mask_mes

    df_filtrado = df_base[mask_final].copy()
    df_filtrado["Fecha_Entrega"] = fecha_parsed[mask_final].dt.normalize()

    info = {
        "filas_antes": int(len(df_base)),
        "validas": int(mask_valida.sum()),
        "excluidas_invalidas": int((~mask_valida).sum()),
        "excluidas_mes": int((mask_valida & ~mask_mes).sum()),
        "filas_despues": int(len(df_filtrado)),
    }
    return df_filtrado, info


def filtrar_acuerdos_vigentes(df_base: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, int] | None]:
    """Excluye acuerdos vigentes de los ultimos 7 dias si las columnas existen."""
    if "Gestion_Estado" not in df_base.columns or "Fecha_Gestion" not in df_base.columns:
        return df_base.copy(), None

    estados_acuerdo = ["07. Promesa de Pago Pactada", "08. Gestión de Refinanciación"]

    df_work = df_base.copy()
    df_work["Gestion_Estado"] = df_work["Gestion_Estado"].astype(str).str.strip()
    df_work["Gestion_Estado"] = df_work["Gestion_Estado"].replace(["nan", "NaN", "None", "NaT"], "")
    df_work["_fecha_gestion_parsed"] = pd.to_datetime(df_work["Fecha_Gestion"], dayfirst=True, errors="coerce")

    fecha_actual = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    df_work["_dias_desde_gestion"] = (fecha_actual - df_work["_fecha_gestion_parsed"]).dt.days

    mask_acuerdo_vigente = (
        df_work["Gestion_Estado"].isin(estados_acuerdo)
        & df_work["_dias_desde_gestion"].notna()
        & (df_work["_dias_desde_gestion"] <= 7)
    )

    df_filtrado = df_work[~mask_acuerdo_vigente].copy()
    df_filtrado = df_filtrado.drop(columns=["_fecha_gestion_parsed", "_dias_desde_gestion"])

    info = {
        "filas_antes": int(len(df_base)),
        "filas_con_estado_acuerdo": int(df_work["Gestion_Estado"].isin(estados_acuerdo).sum()),
        "filas_excluidas": int(mask_acuerdo_vigente.sum()),
        "filas_despues": int(len(df_filtrado)),
    }
    return df_filtrado, info


def consolidar_grupo(grupo: pd.DataFrame) -> pd.Series:
    """Consolida productos por Cliente_BT en una sola fila."""
    resultado = grupo.iloc[0].copy()

    for col in COLUMNAS_NUMERICAS:
        if col in grupo.columns:
            resultado[col] = grupo[col].sum()

    if "Dias_Mora" in grupo.columns:
        resultado["Dias_Mora"] = pd.to_numeric(grupo["Dias_Mora"], errors="coerce").max()

    for col in ["NumeroOperacion", "AgrupadorProducto"]:
        if col in grupo.columns:
            valores = grupo[col].dropna().astype(str)
            valores = valores[valores != ""]
            if col == "AgrupadorProducto":
                valores = valores.apply(corregir_codificacion_texto)
            resultado[col] = ",".join(valores.unique()) if len(valores) > 0 else ""

    return resultado


def deduplicar_por_telefonos(
    df_base: pd.DataFrame,
    cancel_event: Event | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Deduplica por telefonos preservando mayor monto y mas productos."""
    df_work = df_base.copy()

    if "NumeroOperacion" in df_work.columns:
        df_work["_num_productos"] = df_work["NumeroOperacion"].apply(
            lambda x: len(str(x).split(",")) if pd.notna(x) and str(x).strip() != "" else 0
        )
    else:
        df_work["_num_productos"] = 0

    df_work["_monto_numerico"] = df_work["MontoAdeudado"].apply(
        lambda x: float(str(x).replace(",", ".")) if pd.notna(x) and str(x).strip() != "" else 0
    )

    for col in ["NumeroTelefono", "NumeroTrabajo", "NumeroCelular"]:
        if col not in df_work.columns:
            df_work[col] = ""
        df_work[f"_{col}_clean"] = df_work[col].astype(str).str.strip()
        df_work[f"_{col}_clean"] = df_work[f"_{col}_clean"].replace(["nan", "NaN", "None", "NaT", ""], "")

    clientes_unicos = []
    clientes_descartados = []
    procesados = set()

    for index, (_, fila) in enumerate(df_work.iterrows()):
        if index % 200 == 0 and cancel_event is not None and cancel_event.is_set():
            raise PipelineCancelledError("Cancelado por usuario")

        cliente_bt = fila["Cliente_BT"]
        if cliente_bt in procesados:
            continue

        tel = fila.get("_NumeroTelefono_clean", "")
        trab = fila.get("_NumeroTrabajo_clean", "")
        cel = fila.get("_NumeroCelular_clean", "")

        if tel == "" and trab == "" and cel == "":
            clientes_unicos.append(fila)
            procesados.add(cliente_bt)
            continue

        masks = []
        if tel:
            masks.append(df_work["_NumeroTelefono_clean"] == tel)
        if trab:
            masks.append(df_work["_NumeroTrabajo_clean"] == trab)
        if cel:
            masks.append(df_work["_NumeroCelular_clean"] == cel)

        mask_final = masks[0]
        for mask in masks[1:]:
            mask_final = mask_final | mask

        grupo = df_work[mask_final].copy()
        if len(grupo) == 1:
            clientes_unicos.append(fila)
            procesados.add(cliente_bt)
            continue

        grupo_ordenado = grupo.sort_values(by=["_monto_numerico", "_num_productos"], ascending=[False, False])
        mejor = grupo_ordenado.iloc[0]
        clientes_unicos.append(mejor)

        for _, descartado in grupo_ordenado.iloc[1:].iterrows():
            if descartado["Cliente_BT"] not in procesados:
                clientes_descartados.append(descartado)

        for cliente in grupo["Cliente_BT"]:
            procesados.add(cliente)

    df_unicos = pd.DataFrame(clientes_unicos)
    df_descartados = pd.DataFrame(clientes_descartados) if clientes_descartados else pd.DataFrame()

    if "_orden_original" in df_unicos.columns:
        df_unicos = df_unicos.sort_values(by="_orden_original", kind="stable").reset_index(drop=True)
    if not df_descartados.empty and "_orden_original" in df_descartados.columns:
        df_descartados = df_descartados.sort_values(by="_orden_original", kind="stable").reset_index(drop=True)

    columnas_aux = [
        "_num_productos",
        "_monto_numerico",
        "_NumeroTelefono_clean",
        "_NumeroTrabajo_clean",
        "_NumeroCelular_clean",
    ]
    for col in columnas_aux:
        if col in df_unicos.columns:
            df_unicos = df_unicos.drop(columns=[col])
        if not df_descartados.empty and col in df_descartados.columns:
            df_descartados = df_descartados.drop(columns=[col])

    return df_unicos, df_descartados


def deduplicar_por_telefonos_back_base(df_base: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Usa la deduplicacion exacta de back-base para asegurar paridad."""
    module = _load_back_base_generator_module()
    if module is None:
        return deduplicar_por_telefonos(df_base)

    if hasattr(module, "deduplicar_por_telefonos"):
        return module.deduplicar_por_telefonos(df_base)
    return deduplicar_por_telefonos(df_base)


def _load_back_base_generator_module():
    try:
        base_generator_path = Path(__file__).resolve().parents[2] / "back-base" / "procesos" / "base_generator.py"
        if not base_generator_path.exists():
            return None
        spec = importlib.util.spec_from_file_location("back_base_generator", base_generator_path)
        if spec is None or spec.loader is None:
            return None
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module
    except Exception:
        return None


def _normalizar_texto_vacio(valor: object) -> str:
    if pd.isna(valor):
        return ""

    texto = str(valor).strip()
    if texto.lower() in {"", "nan", "none", "nat"}:
        return ""
    return texto


def filtrar_filas_sin_telefono(df_base: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, int] | None]:
    """Excluye filas donde NumeroCelular y NumeroTelefono estan vacios."""
    if "NumeroCelular" not in df_base.columns or "NumeroTelefono" not in df_base.columns:
        return df_base.copy(), None

    df_work = df_base.copy()
    cel_clean = df_work["NumeroCelular"].apply(_normalizar_texto_vacio)
    tel_clean = df_work["NumeroTelefono"].apply(_normalizar_texto_vacio)

    mask_sin_telefono = (cel_clean == "") & (tel_clean == "")
    df_filtrado = df_work[~mask_sin_telefono].copy()

    info = {
        "filas_antes": int(len(df_base)),
        "filas_excluidas": int(mask_sin_telefono.sum()),
        "filas_despues": int(len(df_filtrado)),
    }
    return df_filtrado, info


def _guardar_xlsx(df_salida: pd.DataFrame, ruta_salida: Path) -> None:
    df_salida.to_excel(ruta_salida, index=False)

    wb = load_workbook(ruta_salida)
    ws = wb.active

    if "Fecha_Entrega" in list(df_salida.columns):
        indice = list(df_salida.columns).index("Fecha_Entrega") + 1
        ws.column_dimensions[ws.cell(row=1, column=indice).column_letter].width = 14
        for row_idx in range(2, ws.max_row + 1):
            celda = ws.cell(row=row_idx, column=indice)
            if isinstance(celda.value, (datetime, date)):
                celda.number_format = "DD-MMM-YY"

    wb.save(ruta_salida)


def ejecutar_pipeline_wfm(
    path_entrada: Path,
    meses_permitidos: list[int],
    incluir_fechas_vacias: bool = False,
    progress_callback: ProgressCallback | None = None,
    cancel_event: Event | None = None,
) -> dict[str, Any]:
    """Ejecuta pipeline completo para un archivo y retorna resultado estructurado."""
    logs: list[str] = []
    artifacts: list[dict[str, Any]] = []
    path_salida: Path | None = None
    path_copia_entrada: Path | None = None
    progreso_actual = _emitir_progreso(progress_callback, 0, "Preparando procesamiento")

    try:
        _cancelar_si_corresponde(cancel_event, logs, progress_callback, progreso_actual)

        if not path_entrada.exists():
            raise ValueError(f"Archivo no encontrado: {path_entrada}")

        if path_entrada.suffix.lower() not in {".xlsx", ".xls", ".csv"}:
            raise ValueError("Formato no soportado. Usar .xlsx, .xls o .csv")

        if not meses_permitidos:
            raise ValueError("Debe seleccionar al menos un mes para Fecha_Entrega.")

        run_context = _build_run_context(datetime.now())

        base_dir = obtener_carpeta_base()
        _, carpeta_entrada, carpeta_salida = _crear_estructura_historial(base_dir, run_context["fecha_carpeta"])
        timestamp = run_context["timestamp"]

        path_copia_entrada = _copiar_entrada_historial(path_entrada, carpeta_entrada, timestamp)
        logs.append(f"Entrada copiada a historial: {path_copia_entrada}")
        progreso_actual = _emitir_progreso(progress_callback, 10, "Entrada copiada a historial")
        _cancelar_si_corresponde(cancel_event, logs, progress_callback, progreso_actual)

        df_base, detalle_lectura = leer_archivo_entrada(path_copia_entrada)
        df_base = _limpiar_nombres_columnas(df_base)
        logs.append(f"Lectura OK: {detalle_lectura}")
        logs.append(f"Filas leidas: {len(df_base)}")
        progreso_actual = _emitir_progreso(progress_callback, 22, "Archivo leido y validado")
        _cancelar_si_corresponde(cancel_event, logs, progress_callback, progreso_actual)

        faltantes = sorted(COLUMNAS_REQUERIDAS - set(df_base.columns))
        if faltantes:
            raise ValueError(f"Faltan columnas requeridas: {', '.join(faltantes)}")

        df_base = _normalizar_columnas_numericas(df_base)
        df_base = _normalizar_columnas_texto(df_base)
        df_base = _normalizar_columnas_telefono(df_base)
        logs.append("Normalizacion telefonos: prefijos 54/549 + remocion 15/0 aplicada")
        progreso_actual = _emitir_progreso(progress_callback, 35, "Normalizando datos y telefonos")
        _cancelar_si_corresponde(cancel_event, logs, progress_callback, progreso_actual)

        if "EstadoDescripcion" in df_base.columns:
            antes = len(df_base)
            df_base["EstadoDescripcion"] = df_base["EstadoDescripcion"].astype(str).str.strip()
            df_base = df_base[df_base["EstadoDescripcion"] != "Cancelada"].copy()
            logs.append(f"Filtro EstadoDescripcion != Cancelada: {antes} -> {len(df_base)}")

        if "SaldoCapital" in df_base.columns:
            antes = len(df_base)
            mask_saldo_cero = df_base["SaldoCapital"].notna() & (df_base["SaldoCapital"] == 0)
            df_base = df_base[~mask_saldo_cero].copy()
            logs.append(f"Filtro SaldoCapital != 0: {antes} -> {len(df_base)}")

        df_base = df_base[df_base["MontoAdeudado"].notna() & (df_base["MontoAdeudado"] > 0)].copy()
        logs.append(f"Filtro MontoAdeudado > 0: filas restantes {len(df_base)}")
        if df_base.empty:
            raise ValueError("No quedaron filas luego del filtro de MontoAdeudado > 0.")

        df_base, info_fecha = filtrar_fecha_entrega_por_meses(
            df_base,
            meses_permitidos,
            incluir_fechas_vacias=incluir_fechas_vacias,
        )
        logs.append(
            "Filtro Fecha_Entrega: "
            f"validas={info_fecha['validas']}, "
            f"invalidas={info_fecha['excluidas_invalidas']}, "
            f"mes_fuera={info_fecha['excluidas_mes']}, "
            f"resultado={info_fecha['filas_despues']}"
        )
        if df_base.empty:
            raise ValueError("No quedaron filas luego del filtro de Fecha_Entrega.")
        progreso_actual = _emitir_progreso(progress_callback, 55, "Filtros principales aplicados")
        _cancelar_si_corresponde(cancel_event, logs, progress_callback, progreso_actual)

        df_base, info_acuerdos = filtrar_acuerdos_vigentes(df_base)
        if info_acuerdos is None:
            logs.append("Filtro acuerdos vigentes omitido: columnas no disponibles.")
        else:
            logs.append(
                "Filtro acuerdos vigentes: "
                f"con_estado={info_acuerdos['filas_con_estado_acuerdo']}, "
                f"excluidas={info_acuerdos['filas_excluidas']}, "
                f"resultado={info_acuerdos['filas_despues']}"
            )
        if df_base.empty:
            raise ValueError("No quedaron filas luego del filtro de acuerdos vigentes.")

        # ROMAN y E1KIA deben respetar exactamente la seleccion aplicada en UI
        df_para_auxiliares = df_base.copy()

        if "Cliente_BT" not in df_base.columns:
            raise ValueError("Falta columna obligatoria 'Cliente_BT' para consolidar.")

        columnas_originales = list(df_base.columns)
        df_consolidado = (
            df_base.groupby("Cliente_BT", sort=False)
            .apply(consolidar_grupo)
            .reset_index(drop=True)
        )
        logs.append(f"Consolidacion por Cliente_BT: {len(df_base)} -> {len(df_consolidado)}")

        df_unicos, df_descartados = deduplicar_por_telefonos(df_consolidado, cancel_event=cancel_event)
        logs.append(f"Deduplicacion telefonos: resultado={len(df_unicos)}, descartados={len(df_descartados)}")
        progreso_actual = _emitir_progreso(progress_callback, 70, "Consolidando y deduplicando")
        _cancelar_si_corresponde(cancel_event, logs, progress_callback, progreso_actual)

        df_unicos, info_telefonos = filtrar_filas_sin_telefono(df_unicos)
        if info_telefonos is None:
            logs.append("Filtro filas sin telefono omitido: columnas no disponibles.")
        else:
            logs.append(
                "Filtro filas sin telefono: "
                f"excluidas={info_telefonos['filas_excluidas']}, "
                f"resultado={info_telefonos['filas_despues']}"
            )

        df_salida = df_unicos[[c for c in columnas_originales if c in df_unicos.columns]].copy()
        fecha_nombre = run_context["fecha_ddmmyyyy"]
        nombre_salida = f"base_recibida_BANCOR_conFiltros_{fecha_nombre}_{timestamp}.xlsx"
        path_salida = carpeta_salida / nombre_salida
        _cancelar_si_corresponde(cancel_event, logs, progress_callback, progreso_actual)
        _guardar_xlsx(df_salida, path_salida)
        progreso_actual = _emitir_progreso(progress_callback, 84, "Archivo principal XLSX generado")
        _cancelar_si_corresponde(cancel_event, logs, progress_callback, progreso_actual)

        artifacts.append(_crear_artifact("xlsx", nombre_salida, path_salida, "generated"))

        artifacts_auxiliares = _exportar_auxiliares(
            df_para_auxiliares,
            carpeta_salida,
            run_context["fecha_yyyymmdd"],
            cancel_event=cancel_event,
            force_back_base_sync=False,
        )
        artifacts.extend(artifacts_auxiliares)

        artifact_roman = next((a for a in artifacts_auxiliares if a.get("name") == "roman"), None)
        if isinstance(artifact_roman, dict) and artifact_roman.get("status") == "generated":
            path_roman = Path(str(artifact_roman.get("path", "")))
            if path_roman.exists():
                logs.extend(_validar_variables_roman_obligatorias(path_roman))
            else:
                logs.append("ADVERTENCIA ROMAN: no se encontro el archivo ROMAN para validar variables.")
        else:
            logs.append("ADVERTENCIA ROMAN: validacion de variables omitida porque ROMAN no se genero.")
        progreso_actual = _emitir_progreso(progress_callback, 95, "Exportes auxiliares finalizados")
        _cancelar_si_corresponde(cancel_event, logs, progress_callback, progreso_actual)

        for artifact in artifacts_auxiliares:
            if artifact["status"] == "generated":
                logs.append(f"[{artifact['name']}] generated: {artifact['path']}")
            else:
                logs.append(f"[{artifact['name']}] failed: {artifact.get('error') or 'Error desconocido'}")

        status = _determinar_status_corrida(artifacts)
        logs.append(f"Estado final de corrida: {status}")
        mensaje_final = (
            "Procesamiento completado"
            if status == "success"
            else "Procesamiento completado con advertencias"
        )
        progreso_actual = _emitir_progreso(progress_callback, 100, mensaje_final)

        return {
            "ok": status == "success",
            "status": status,
            "logs": logs,
            "output_path": str(path_salida),
            "input_history_path": str(path_copia_entrada),
            "rows_output": int(len(df_salida)),
            "artifacts": artifacts,
        }
    except PipelineCancelledError as exc:
        return {
            "ok": False,
            "status": "cancelled",
            "logs": logs,
            "error": str(exc),
            "output_path": str(path_salida) if path_salida else "",
            "input_history_path": str(path_copia_entrada) if path_copia_entrada else "",
            "artifacts": artifacts,
        }
    except Exception as exc:
        logs.append(f"ERROR: {exc}")
        _emitir_progreso(progress_callback, progreso_actual, f"Error: {exc}")
        return {
            "ok": False,
            "status": "failed",
            "logs": logs,
            "error": str(exc),
            "output_path": str(path_salida) if path_salida else "",
            "input_history_path": str(path_copia_entrada) if path_copia_entrada else "",
            "artifacts": artifacts,
        }


if __name__ == "__main__":
    print("Modulo de pipeline WFM. Abrir UI con: python filtrosAplicados_base_BANCOR/main.py")

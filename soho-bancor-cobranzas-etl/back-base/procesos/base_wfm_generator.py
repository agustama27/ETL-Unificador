"""
base_wfm_generator.py
---------------------
Procesa archivos de la carpeta back-base/base_WFM/ (CSV o XLSX) con los
siguientes filtros:
  1. MontoAdeudado > 0 (excluir clientes sin deuda)
  2. Fecha_Entrega en Febrero o Marzo (meses 2 y 3)
  3. Excluir acuerdos vigentes: Gestion_Estado en lista de acuerdos
      Y Fecha_Gestion dentro de los últimos 7 días.
  4. Deduplicación post-consolidación por teléfono (NumeroTelefono,
      NumeroTrabajo, NumeroCelular): mantener el de mayor MontoAdeudado;
      tiebreak = mayor cantidad de productos (NumeroOperacion).

Salida:
  - back-base/base_WFM/resultado/base_wfm_DDMMAAAA.csv
  - back-base/base_WFM/resultado/telefonos_x_cliente_DDMMAAAA.csv
  - back-base/base_WFM/resultado/descartados_por_telefono_DDMMAAAA.csv

Standalone: ejecutar con
    python back-base/procesos/base_wfm_generator.py
desde la raíz del proyecto.
"""

import sys
import os
import re
from pathlib import Path
from datetime import datetime, timedelta, date

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Helpers reutilizados del pipeline estándar (base_generator.py)
# ---------------------------------------------------------------------------

def corregir_codificacion_texto(texto: str) -> str:
    """
    Corrige problemas de codificación (mojibake), p.ej. 'CrÃ©dito' -> 'Crédito'.
    Ocurre cuando un archivo UTF-8 fue leído como latin-1.
    """
    if pd.isna(texto) or texto == '':
        return texto if isinstance(texto, str) else ''

    texto_str = str(texto)

    if 'Ã' in texto_str:
        try:
            return texto_str.encode('latin-1').decode('utf-8')
        except (UnicodeEncodeError, UnicodeDecodeError):
            texto_corregido = texto_str
            texto_corregido = texto_corregido.replace('Ã¡', 'á').replace('Ã©', 'é').replace('Ã­', 'í')
            texto_corregido = texto_corregido.replace('Ã³', 'ó').replace('Ãº', 'ú').replace('Ã±', 'ñ')
            texto_corregido = texto_corregido.replace('Ã', 'Á').replace('Ã‰', 'É').replace('Ã', 'Í')
            texto_corregido = texto_corregido.replace('Ã"', 'Ó').replace('Ãš', 'Ú')
            texto_corregido = texto_corregido.replace("Ã'", "Ñ")
            texto_corregido = texto_corregido.replace('Ã¼', 'ü').replace('Ãœ', 'Ü')
            return texto_corregido

    return texto_str


def deduplicar_por_telefonos(df):
    """
    Deduplica clientes que comparten números de teléfono.
    Mantiene el cliente con mayor MontoAdeudado.
    Tiebreak: mayor cantidad de productos (NumeroOperacion).

    Returns:
        Tuple (df_unicos, df_descartados)
    """
    df = df.copy()

    df['_num_productos'] = df['NumeroOperacion'].apply(
        lambda x: len(str(x).split(',')) if pd.notna(x) and str(x).strip() != '' else 0
    )

    df['_monto_numerico'] = df['MontoAdeudado'].apply(
        lambda x: float(str(x).replace(',', '.')) if pd.notna(x) and str(x).strip() != '' else 0
    )

    clientes_unicos = []
    clientes_descartados = []
    procesados = set()

    for col in ['NumeroTelefono', 'NumeroTrabajo', 'NumeroCelular']:
        if col in df.columns:
            df[f'_{col}_clean'] = df[col].astype(str).str.strip()
            df[f'_{col}_clean'] = df[f'_{col}_clean'].replace(['nan', 'NaN', 'None', 'NaT', ''], '')

    for idx, fila in df.iterrows():
        cliente_bt = fila['Cliente_BT']

        if cliente_bt in procesados:
            continue

        tel = fila.get('_NumeroTelefono_clean', '')
        trab = fila.get('_NumeroTrabajo_clean', '')
        cel = fila.get('_NumeroCelular_clean', '')

        if tel == '' and trab == '' and cel == '':
            clientes_unicos.append(fila)
            procesados.add(cliente_bt)
            continue

        masks = []
        if tel != '':
            masks.append(df['_NumeroTelefono_clean'] == tel)
        if trab != '':
            masks.append(df['_NumeroTrabajo_clean'] == trab)
        if cel != '':
            masks.append(df['_NumeroCelular_clean'] == cel)

        if masks:
            mask_final = masks[0]
            for m in masks[1:]:
                mask_final = mask_final | m
        else:
            clientes_unicos.append(fila)
            procesados.add(cliente_bt)
            continue

        grupo = df[mask_final].copy()

        if len(grupo) == 1:
            clientes_unicos.append(fila)
            procesados.add(cliente_bt)
        else:
            grupo_ordenado = grupo.sort_values(
                by=['_monto_numerico', '_num_productos'],
                ascending=[False, False]
            )
            mejor_cliente = grupo_ordenado.iloc[0]
            clientes_unicos.append(mejor_cliente)

            for _, cliente_descartado in grupo_ordenado.iloc[1:].iterrows():
                if cliente_descartado['Cliente_BT'] not in procesados:
                    clientes_descartados.append(cliente_descartado)

            for cliente_bt_proc in grupo['Cliente_BT']:
                procesados.add(cliente_bt_proc)

    df_unicos = pd.DataFrame(clientes_unicos)
    df_descartados = pd.DataFrame(clientes_descartados) if clientes_descartados else pd.DataFrame()

    cols_aux = [
        '_num_productos', '_monto_numerico',
        '_NumeroTelefono_clean', '_NumeroTrabajo_clean', '_NumeroCelular_clean'
    ]
    for col in cols_aux:
        if col in df_unicos.columns:
            df_unicos = df_unicos.drop(columns=[col])
        if len(df_descartados) > 0 and col in df_descartados.columns:
            df_descartados = df_descartados.drop(columns=[col])

    return df_unicos, df_descartados


def filtrar_acuerdos_vigentes(df):
    """
    Excluye filas con acuerdos vigentes (<= 7 días de antigüedad).
    Si Fecha_Gestion está vacía/inválida, la fila se MANTIENE.
    """
    if 'Gestion_Estado' not in df.columns or 'Fecha_Gestion' not in df.columns:
        print("  Advertencia: columnas Gestion_Estado/Fecha_Gestion no encontradas. "
              "Filtro de acuerdos vigentes omitido.")
        return df

    filas_antes = len(df)

    estados_acuerdo = [
        '07. Promesa de Pago Pactada',
        '08. Gestión de Refinanciación',
    ]

    # Limpiar Gestion_Estado
    df = df.copy()
    df['Gestion_Estado'] = df['Gestion_Estado'].astype(str).str.strip()
    df['Gestion_Estado'] = df['Gestion_Estado'].replace(['nan', 'NaN', 'None', 'NaT'], '')

    # Parsear Fecha_Gestion (puede venir como datetime desde Excel o string desde CSV)
    df['_fecha_gestion_parsed'] = pd.to_datetime(df['Fecha_Gestion'], dayfirst=True, errors='coerce')

    fecha_actual = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    df['_dias_desde_gestion'] = (fecha_actual - df['_fecha_gestion_parsed']).dt.days

    mask_acuerdo_vigente = (
        (df['Gestion_Estado'].isin(estados_acuerdo)) &
        (df['_dias_desde_gestion'].notna()) &
        (df['_dias_desde_gestion'] <= 7)
    )

    filas_excluidas = mask_acuerdo_vigente.sum()
    print(f"\n  --- Filtro de acuerdos vigentes ---")
    print(f"  Filas con estado de acuerdo: {df['Gestion_Estado'].isin(estados_acuerdo).sum()}")
    print(f"  Filas con acuerdo vigente (<= 7 días): {filas_excluidas}")

    if filas_excluidas > 0:
        df_excluidos = df[mask_acuerdo_vigente]
        for estado in estados_acuerdo:
            count = (df_excluidos['Gestion_Estado'] == estado).sum()
            if count > 0:
                print(f"    - {estado}: {count} filas excluidas")

    df_filtrado = df[~mask_acuerdo_vigente].copy()
    df_filtrado = df_filtrado.drop(columns=['_fecha_gestion_parsed', '_dias_desde_gestion'])

    print(f"  Filas antes del filtro: {filas_antes}")
    print(f"  Filas después del filtro: {len(df_filtrado)}")

    return df_filtrado


def filtrar_fecha_entrega_febrero_marzo(df):
    """
    Conserva solo filas con Fecha_Entrega parseable en meses Febrero/Marzo.
    - Admite fechas típicas del flujo (dd/mm/yyyy, yyyy-mm-dd, datetime).
    - Intenta parseo adicional para seriales de Excel.
    - Nulos o parse inválido se excluyen.
    """
    if 'Fecha_Entrega' not in df.columns:
        print("  ERROR: columna Fecha_Entrega no encontrada. Saltando archivo.")
        return pd.DataFrame()

    filas_antes = len(df)
    df = df.copy()

    mapa_meses = {
        'ene': '01', 'enero': '01',
        'feb': '02', 'febrero': '02',
        'mar': '03', 'marzo': '03',
        'abr': '04', 'abril': '04',
        'may': '05', 'mayo': '05',
        'jun': '06', 'junio': '06',
        'jul': '07', 'julio': '07',
        'ago': '08', 'agosto': '08',
        'sep': '09', 'sept': '09', 'septiembre': '09',
        'oct': '10', 'octubre': '10',
        'nov': '11', 'noviembre': '11',
        'dic': '12', 'diciembre': '12',
    }

    def parsear_fecha_entrega(valor):
        if pd.isna(valor):
            return pd.NaT

        if isinstance(valor, (pd.Timestamp, datetime, date)):
            return pd.Timestamp(valor)

        if isinstance(valor, (int, float)) and not isinstance(valor, bool):
            serial = float(valor)
            return pd.to_datetime(pd.Series([serial]), unit='D', origin='1899-12-30', errors='coerce').iloc[0]

        texto = str(valor).strip()
        if texto in ('', 'nan', 'NaN', 'None', 'NaT'):
            return pd.NaT

        texto_lower = texto.lower()

        if re.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}', texto_lower):
            return pd.to_datetime(texto, errors='coerce')

        try:
            serial = float(texto.replace(',', '.'))
            return pd.to_datetime(pd.Series([serial]), unit='D', origin='1899-12-30', errors='coerce').iloc[0]
        except ValueError:
            pass

        texto_normalizado = texto_lower
        for mes, nro in mapa_meses.items():
            texto_normalizado = re.sub(rf'\b{mes}\b', nro, texto_normalizado)

        return pd.to_datetime(texto_normalizado, dayfirst=True, errors='coerce')

    fecha_parsed = df['Fecha_Entrega'].apply(parsear_fecha_entrega)

    mask_valida = fecha_parsed.notna()
    mask_mes_permitido = fecha_parsed.dt.month.isin([2, 3])
    mask_final = mask_valida & mask_mes_permitido

    excluidas_invalidas = int((~mask_valida).sum())
    excluidas_mes = int((mask_valida & ~mask_mes_permitido).sum())

    df_filtrado = df[mask_final].copy()
    df_filtrado['Fecha_Entrega'] = fecha_parsed[mask_final].dt.normalize()

    print("\n  --- Filtro Fecha_Entrega (Febrero/Marzo) ---")
    print(f"  Filas con Fecha_Entrega válida: {int(mask_valida.sum())}")
    print(f"  Excluidas por fecha nula/inválida: {excluidas_invalidas}")
    print(f"  Excluidas por mes fuera de 2/3: {excluidas_mes}")
    print(f"  Filas antes del filtro: {filas_antes}")
    print(f"  Filas después del filtro: {len(df_filtrado)}")

    return df_filtrado


def leer_archivo_wfm(ruta):
    """
    Lee un archivo WFM (XLSX o CSV).
    Para CSV prueba encodings: latin-1, iso-8859-1, cp1252, utf-8.
    Retorna (DataFrame, encoding_usado).
    """
    ruta = Path(ruta)
    ext = ruta.suffix.lower()

    if ext in ('.xlsx', '.xls'):
        print(f"  Leyendo Excel: {ruta.name}")
        df = pd.read_excel(ruta)
        return df, 'xlsx'

    # CSV
    codificaciones = ['latin-1', 'iso-8859-1', 'cp1252', 'utf-8', 'utf-16']
    for enc in codificaciones:
        try:
            df = pd.read_csv(ruta, sep=';', encoding=enc, low_memory=False)
            print(f"  Leyendo CSV con encoding={enc}: {ruta.name}")
            return df, enc
        except UnicodeDecodeError:
            continue
        except Exception:
            raise

    print("  Advertencia: intentando con latin-1 + errors=replace")
    df = pd.read_csv(ruta, sep=';', encoding='latin-1', errors='replace', low_memory=False)
    return df, 'latin-1'


def limpiar_nombre_columnas(df):
    """
    Corrige mojibake en nombres de columnas.
    Ej: 'GestionDescripciÃ³n' -> 'GestionDescripción'
    """
    rename_map = {}
    for col in df.columns:
        col_fixed = corregir_codificacion_texto(col)
        if col_fixed != col:
            rename_map[col] = col_fixed
    if rename_map:
        print(f"  Columnas renombradas por encoding: {rename_map}")
        df = df.rename(columns=rename_map)
    return df


def consolidar_grupo(grupo):
    """
    Consolida múltiples productos de un mismo Cliente_BT en 1 fila.
    - MontoAdeudado, MontoVencido, SaldoCapital, InteresAdeudado,
      IVAInteresAdeudado, OFERTA_Importe: suma
    - Dias_Mora: máximo
    - NumeroOperacion, AgrupadorProducto: concatenar únicos con coma
    - Resto de columnas: primera fila
    """
    resultado = grupo.iloc[0].copy()

    # Columnas a sumar
    cols_suma = [
        'MontoAdeudado', 'MontoVencido', 'SaldoCapital',
        'InteresAdeudado', 'IVAInteresAdeudado', 'OFERTA_Importe',
        'Deuda_vencida_Clte', 'CapitalOriginal',
    ]
    for col in cols_suma:
        if col in grupo.columns:
            resultado[col] = grupo[col].sum()

    if 'Dias_Mora' in grupo.columns:
        resultado['Dias_Mora'] = grupo['Dias_Mora'].max()

    for col in ['NumeroOperacion', 'AgrupadorProducto']:
        if col in grupo.columns:
            valores = grupo[col].dropna().astype(str)
            valores = valores[valores != '']
            if col == 'AgrupadorProducto':
                valores = valores.apply(corregir_codificacion_texto)
            resultado[col] = ','.join(valores.unique()) if len(valores) > 0 else ''

    return resultado


def formato_europeo(valor):
    """Convierte float a string con coma decimal (formato europeo)."""
    if pd.isna(valor):
        return ''
    s = str(valor)
    if '.' in s:
        return s.replace('.', ',')
    return s


def procesar_base_wfm():
    """
    Función principal. Lee archivos de back-base/base_WFM/,
    aplica filtros, consolida y guarda resultados.
    """
    # -----------------------------------------------------------------------
    # Rutas
    # -----------------------------------------------------------------------
    base_dir = Path(__file__).parent.parent
    carpeta_entrada = base_dir / "base_WFM"
    carpeta_salida = carpeta_entrada / "resultado"
    carpeta_salida.mkdir(parents=True, exist_ok=True)

    fecha_hoy = datetime.now().strftime('%d%m%Y')

    # -----------------------------------------------------------------------
    # Buscar archivos de entrada (XLSX + CSV)
    # -----------------------------------------------------------------------
    archivos = (
        list(carpeta_entrada.glob("*.xlsx")) +
        list(carpeta_entrada.glob("*.xls")) +
        list(carpeta_entrada.glob("*.csv"))
    )

    if not archivos:
        print("No se encontraron archivos CSV/XLSX en back-base/base_WFM/")
        return

    print(f"\n{'='*70}")
    print("PROCESAMIENTO BASE WFM")
    print(f"{'='*70}")
    print(f"Archivos encontrados: {[a.name for a in archivos]}")

    for archivo in archivos:
        print(f"\n{'-'*60}")
        print(f"Procesando: {archivo.name}")
        print(f"{'-'*60}")

        try:
            # -------------------------------------------------------------------
            # Lectura
            # -------------------------------------------------------------------
            df, encoding_usado = leer_archivo_wfm(archivo)
            total_leidas = len(df)
            print(f"\n  Filas leídas: {total_leidas}")
            print(f"  Columnas: {len(df.columns)}")

            # Corregir nombres de columnas con mojibake
            df = limpiar_nombre_columnas(df)

            # -------------------------------------------------------------------
            # Normalizar columnas numéricas que puedan venir como string (CSV)
            # -------------------------------------------------------------------
            for col_num in ['MontoAdeudado', 'MontoVencido', 'SaldoCapital',
                             'InteresAdeudado', 'IVAInteresAdeudado',
                             'OFERTA_Importe', 'Deuda_vencida_Clte', 'CapitalOriginal']:
                if col_num in df.columns:
                    if df[col_num].dtype == object:
                        df[col_num] = (
                            df[col_num].astype(str)
                            .str.replace(',', '.', regex=False)
                        )
                        df[col_num] = pd.to_numeric(df[col_num], errors='coerce')

            # -------------------------------------------------------------------
            # Filtro previo: EstadoDescripcion != 'Cancelada'
            # -------------------------------------------------------------------
            if 'EstadoDescripcion' in df.columns:
                df['EstadoDescripcion'] = df['EstadoDescripcion'].astype(str).str.strip()
                canceladas = int((df['EstadoDescripcion'] == 'Cancelada').sum())
                df = df[df['EstadoDescripcion'] != 'Cancelada'].copy()
                print(f"  Filtro EstadoDescripcion != 'Cancelada': excluidas {canceladas}, quedan {len(df)}")

            # -------------------------------------------------------------------
            # Filtro previo: SaldoCapital != 0
            # -------------------------------------------------------------------
            if 'SaldoCapital' in df.columns:
                saldo_num = pd.to_numeric(
                    df['SaldoCapital'].astype(str).str.replace(',', '.', regex=False),
                    errors='coerce',
                )
                mask_saldo_cero = saldo_num.notna() & (saldo_num == 0)
                excluidos_saldo = int(mask_saldo_cero.sum())
                df = df[~mask_saldo_cero].copy()
                print(f"  Filtro SaldoCapital != 0: excluidas {excluidos_saldo}, quedan {len(df)}")

            # -------------------------------------------------------------------
            # Filtro 1: MontoAdeudado > 0
            # -------------------------------------------------------------------
            if 'MontoAdeudado' not in df.columns:
                print("  ERROR: columna MontoAdeudado no encontrada. Saltando archivo.")
                continue

            # Asegurar numérico (por si es object en CSV)
            if df['MontoAdeudado'].dtype == object:
                df['MontoAdeudado'] = (
                    df['MontoAdeudado'].astype(str)
                    .str.replace(',', '.', regex=False)
                )
                df['MontoAdeudado'] = pd.to_numeric(df['MontoAdeudado'], errors='coerce')

            df_f1 = df[(df['MontoAdeudado'].notna()) & (df['MontoAdeudado'] > 0)].copy()
            print(f"\n  Filtro 1 (MontoAdeudado > 0):")
            print(f"    Antes: {total_leidas} | Después: {len(df_f1)} | Excluidas: {total_leidas - len(df_f1)}")

            if len(df_f1) == 0:
                print("  No quedan filas tras Filtro 1. Saltando archivo.")
                continue

            # -------------------------------------------------------------------
            # Limpieza de columnas auxiliares antes del Filtro 2
            # -------------------------------------------------------------------
            # Convertir ModuloCodigo a string si existe
            if 'ModuloCodigo' in df_f1.columns:
                df_f1['ModuloCodigo'] = df_f1['ModuloCodigo'].astype(str)

            # NumeroDocumento y Nro Cuenta: enteros sin decimales
            for campo in ['NumeroDocumento', 'Nro Cuenta']:
                if campo in df_f1.columns:
                    df_f1[campo] = df_f1[campo].astype(str)
                    df_f1[campo] = df_f1[campo].replace(['nan', 'NaN', 'None', 'NaT'], '')
                    df_f1[campo] = pd.to_numeric(df_f1[campo], errors='coerce')
                    df_f1[campo] = df_f1[campo].apply(
                        lambda x: '' if pd.isna(x) else str(int(x))
                    )

            # Teléfonos: quitar guiones y número ficticio
            numero_ficticio = '3519999999'
            for campo in ['NumeroTelefono', 'NumeroTrabajo', 'NumeroCelular']:
                if campo in df_f1.columns:
                    df_f1[campo] = df_f1[campo].astype(str)
                    df_f1[campo] = df_f1[campo].replace(['nan', 'NaN', 'None', 'NaT'], '')
                    df_f1[campo] = df_f1[campo].str.replace('-', '', regex=False)
                    reemplazados = (df_f1[campo] == numero_ficticio).sum()
                    if reemplazados > 0:
                        print(f"  Teléfono ficticio '{numero_ficticio}' reemplazado por vacío en {campo}: {reemplazados}")
                    df_f1[campo] = df_f1[campo].replace(numero_ficticio, '')

            # Columnas de texto con posible mojibake
            for campo in ['Estado Cuenta', 'Tasa_40', 'AgrupadorProducto', 'GestionDescripción']:
                if campo in df_f1.columns:
                    df_f1[campo] = df_f1[campo].astype(str)
                    df_f1[campo] = df_f1[campo].replace(['nan', 'NaN', 'None', 'NaT'], '')
                    df_f1[campo] = df_f1[campo].apply(corregir_codificacion_texto)

            # -------------------------------------------------------------------
            # Filtro 2: Fecha_Entrega en Febrero/Marzo
            # -------------------------------------------------------------------
            df_f2 = filtrar_fecha_entrega_febrero_marzo(df_f1)
            print(f"\n  Después de Filtro 2 (Fecha_Entrega mes 2/3): {len(df_f2)}")

            if len(df_f2) == 0:
                print("  No quedan filas tras Filtro 2. Saltando archivo.")
                continue

            # -------------------------------------------------------------------
            # Filtro 3: Excluir acuerdos vigentes
            # -------------------------------------------------------------------
            df_f3 = filtrar_acuerdos_vigentes(df_f2)
            print(f"\n  Después de Filtro 3 (acuerdos vigentes): {len(df_f3)}")

            if len(df_f3) == 0:
                print("  No quedan filas tras Filtro 3. Saltando archivo.")
                continue

            # -------------------------------------------------------------------
            # Consolidar por Cliente_BT (multiproducto -> 1 fila por cliente)
            # -------------------------------------------------------------------
            print(f"\n  Consolidando por Cliente_BT...")
            print(f"    Filas pre-consolidación: {len(df_f3)}")

            # Convertir Dias_Mora a numérico para el max
            if 'Dias_Mora' in df_f3.columns:
                df_f3['Dias_Mora'] = pd.to_numeric(df_f3['Dias_Mora'], errors='coerce')

            if 'NumeroOperacion' in df_f3.columns:
                df_f3['NumeroOperacion'] = df_f3['NumeroOperacion'].astype(str)
                df_f3['NumeroOperacion'] = df_f3['NumeroOperacion'].replace(['nan', 'NaN', 'None', 'NaT'], '')

            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", FutureWarning)
                df_consolidado = (
                    df_f3.groupby('Cliente_BT', sort=False)
                    .apply(consolidar_grupo)
                    .reset_index(drop=True)
                )
            print(f"    Filas post-consolidación (clientes únicos): {len(df_consolidado)}")

            # -------------------------------------------------------------------
            # Filtro 4: Deduplicación por teléfono post-consolidación
            # -------------------------------------------------------------------
            print(f"\n  --- Deduplicación por teléfonos ---")
            print(f"  Filas antes de deduplicación: {len(df_consolidado)}")

            df_consolidado, df_descartados = deduplicar_por_telefonos(df_consolidado)

            print(f"  Filas después de deduplicación: {len(df_consolidado)}")
            print(f"  Clientes descartados: {len(df_descartados)}")

            # -------------------------------------------------------------------
            # Reordenar columnas como el archivo original (preservar las 84 cols)
            # -------------------------------------------------------------------
            columnas_originales = list(df.columns)
            df_salida = df_consolidado[
                [c for c in columnas_originales if c in df_consolidado.columns]
            ].copy()

            # -------------------------------------------------------------------
            # Guardar resultado principal en XLSX (mismo formato que el original)
            # -------------------------------------------------------------------
            nombre_salida = f"base_wfm_{fecha_hoy}.xlsx"
            ruta_salida = carpeta_salida / nombre_salida
            df_salida.to_excel(ruta_salida, index=False)

            # Evitar '#####' en Excel para Fecha_Entrega (columna C)
            wb = load_workbook(ruta_salida)
            ws = wb.active
            ws.column_dimensions['C'].width = 14
            for row in range(2, ws.max_row + 1):
                celda = ws.cell(row=row, column=3)
                if isinstance(celda.value, (datetime, date)):
                    celda.number_format = 'DD-MMM-YY'
            wb.save(ruta_salida)

            print(f"\n  Archivo principal guardado: {ruta_salida}")
            print(f"  Filas en resultado: {len(df_salida)}")

            # -------------------------------------------------------------------
            # Resumen final
            # -------------------------------------------------------------------
            print(f"\n{'='*70}")
            print("RESUMEN")
            print(f"{'='*70}")
            print(f"  Archivo de entrada        : {archivo.name}")
            print(f"  Filas leídas (input)      : {total_leidas}")
            print(f"  Después Filtro 1 (monto)  : {len(df_f1)}")
            print(f"  Después Filtro 2 (fecha)  : {len(df_f2)}")
            print(f"  Después Filtro 3 (acuerd) : {len(df_f3)}")
            print(f"  Después consolidación     : {len(df_consolidado)}")
            print(f"  Después Filtro 4 (telef)  : {len(df_salida)}")
            print(f"  Archivo generado: {ruta_salida}")
            print(f"{'='*70}\n")

        except Exception as e:
            import traceback
            print(f"\n  ERROR procesando {archivo.name}: {e}")
            traceback.print_exc()
            continue


if __name__ == '__main__':
    # Cambiar al directorio raíz del proyecto si se ejecuta como script
    script_dir = Path(__file__).resolve().parent.parent.parent
    os.chdir(script_dir)
    procesar_base_wfm()

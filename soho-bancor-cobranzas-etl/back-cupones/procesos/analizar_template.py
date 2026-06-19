"""
Script temporal para analizar la estructura del template Excel
y identificar los campos que contiene.
"""
from openpyxl import load_workbook
from pathlib import Path

def analizar_template():
    """Analiza el template Excel y muestra su estructura."""
    template_path = Path(__file__).parent.parent / "template" / "CUPON BANCOR - Template.xlsm"
    
    if not template_path.exists():
        print(f"Error: No se encontró el archivo {template_path}")
        return
    
    print(f"Analizando template: {template_path.name}\n")
    
    try:
        # Cargar el workbook
        wb = load_workbook(template_path, keep_vba=True, data_only=False)
        
        print(f"Número de hojas: {len(wb.sheetnames)}")
        print(f"Nombres de hojas: {wb.sheetnames}\n")
        
        # Analizar la hoja "CARGA INICIAL" que es donde se ingresan los datos
        if "CARGA INICIAL" in wb.sheetnames:
            ws = wb["CARGA INICIAL"]
            print("=" * 80)
            print("HOJA DE ENTRADA DE DATOS: CARGA INICIAL")
            print("=" * 80)
            
            print("\nCAMPOS IDENTIFICADOS EN 'CARGA INICIAL' (donde se ingresan los datos):")
            print("-" * 80)
            
            # Analizar las primeras filas en detalle
            campos_entrada = {}
            
            for row_idx in range(1, min(35, ws.max_row + 1)):
                for col_idx in range(1, min(35, ws.max_column + 1)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        valor = str(cell.value).strip()
                        # Buscar celdas que NO son fórmulas (datos de entrada directos)
                        if cell.data_type != 'f' and valor and valor not in ['#N/A', '#VALUE!']:
                            coord = cell.coordinate
                            # Verificar si la celda adyacente tiene un label
                            # Buscar labels en la misma fila o fila anterior
                            label = None
                            # Buscar en la misma fila hacia la izquierda
                            for c in range(max(1, col_idx - 5), col_idx):
                                label_cell = ws.cell(row=row_idx, column=c)
                                if label_cell.value:
                                    label_val = str(label_cell.value).strip()
                                    if label_val and len(label_val) < 50 and label_val not in ['#N/A', '#VALUE!']:
                                        if any(keyword in label_val.upper() for keyword in ['CID', 'SUCURSAL', 'FECHA', 'CUENTA', 'MÓDULO', 'MONEDA', 'OPERACIÓN', 'TIPO', 'DOC', 'CUIL', 'APELLIDO', 'NOMBRE', 'RAZÓN', 'SOCIAL', 'PAPEL', 'CAPITAL', 'INTERES', 'SEGURO', 'GASTO', 'MORA', 'IVA', 'TOTAL']):
                                            label = label_val
                                            break
                            
                            if label or (row_idx <= 20 and col_idx <= 20):  # Mostrar primeras celdas importantes
                                campos_entrada[coord] = {
                                    'valor': valor[:60],
                                    'fila': row_idx,
                                    'columna': col_idx,
                                    'label': label
                                }
            
            # Mostrar campos organizados
            print("\nCAMPOS DE ENTRADA IDENTIFICADOS:")
            print("-" * 80)
            for coord in sorted(campos_entrada.keys(), key=lambda x: (campos_entrada[x]['fila'], campos_entrada[x]['columna'])):
                info = campos_entrada[coord]
                label_str = f" | Label: {info['label']}" if info['label'] else ""
                print(f"Celda {coord:6} | Fila {info['fila']:2} | Col {info['columna']:2}{label_str}")
                print(f"           Valor: {info['valor']}")
                print()
            
            # Mapeo específico basado en el análisis anterior
            print("\n" + "=" * 80)
            print("MAPEO DE CAMPOS PRINCIPALES (basado en estructura conocida):")
            print("=" * 80)
            
            campos_mapeo = {
                'CID': {'celda': 'D4', 'label': 'CID:', 'ubicacion_label': 'B4'},
                'SUCURSAL': {'celda': 'G4', 'label': 'Sucursal:', 'ubicacion_label': 'G4'},
                'FECHA_PAGO': {'celda': 'N9', 'label': 'FECHA DE PAGO', 'ubicacion_label': 'N8'},
                'CUENTA_CLIENTE': {'celda': 'T9', 'label': 'CUENTA CLIENTE', 'ubicacion_label': 'T8'},
                'MODULO': {'celda': 'B12', 'label': 'MÓDULO', 'ubicacion_label': 'B11'},
                'MONEDA': {'celda': 'N12', 'label': 'MONEDA', 'ubicacion_label': 'N11'},
                'NRO_OPERACION': {'celda': 'B15', 'label': 'N° DE OPERACIÓN', 'ubicacion_label': 'B14'},
                'TIPO_OPERACION': {'celda': 'J15', 'label': 'TIPO OPERACIÓN', 'ubicacion_label': 'J14'},
                'PAPEL': {'celda': 'Q15', 'label': 'PAPEL', 'ubicacion_label': 'Q14'},
                'TIPO_DOC': {'celda': 'B18', 'label': 'TIPO DOC.', 'ubicacion_label': 'B17'},
                'CUIL_CUIT': {'celda': 'G18', 'label': 'N° CUIT/CUIL', 'ubicacion_label': 'G17'},
                'NOMBRE_CLIENTE': {'celda': 'L18', 'label': 'APELLIDO Y NOMBRES / RAZON SOCIAL', 'ubicacion_label': 'L17'},
            }
            
            print("\nCAMPOS EN LA HOJA 'CARGA INICIAL':")
            print("-" * 80)
            for campo, info in campos_mapeo.items():
                celda = ws[info['celda']]
                valor_actual = str(celda.value) if celda.value is not None else "(vacío)"
                es_formula = celda.data_type == 'f'
                tipo = "FÓRMULA" if es_formula else "VALOR"
                print(f"{campo:20} | Celda: {info['celda']:6} | Tipo: {tipo:8} | Label: {info['label']:40} | Valor: {valor_actual[:40]}")
            
            # Buscar campos de montos (probablemente en columnas AE en adelante)
            print("\n" + "-" * 80)
            print("BUSCANDO CAMPOS DE MONTOS (TOTAL PAGADO, CAPITAL, INTERESES, etc.):")
            print("-" * 80)
            
            # Las fórmulas en Cobranzas referencian columnas AE, así que busquemos ahí
            campos_montos = {}
            for row_idx in range(1, 20):
                for col_letter in ['AE', 'AF', 'AG', 'AH', 'AI']:
                    try:
                        col_idx = ord(col_letter[0]) - ord('A') + 1
                        if len(col_letter) == 2:
                            col_idx = (ord(col_letter[0]) - ord('A') + 1) * 26 + (ord(col_letter[1]) - ord('A') + 1)
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            valor = str(cell.value).strip()
                            if valor and valor not in ['#N/A', '#VALUE!']:
                                coord = cell.coordinate
                                campos_montos[coord] = {
                                    'valor': valor,
                                    'fila': row_idx,
                                    'columna': col_idx
                                }
                    except:
                        pass
            
            if campos_montos:
                print("\nCAMPOS DE MONTOS ENCONTRADOS:")
                for coord in sorted(campos_montos.keys(), key=lambda x: (campos_montos[x]['fila'], campos_montos[x]['columna'])):
                    info = campos_montos[coord]
                    print(f"  {coord:6} | Fila {info['fila']:2} | Valor: {info['valor'][:50]}")
            else:
                print("No se encontraron campos de montos en las columnas esperadas.")
                print("Pueden estar en otras ubicaciones o ser calculados por fórmulas.")
        
        print("\n" + "=" * 80)
        print("NOTA IMPORTANTE:")
        print("=" * 80)
        print("Los datos se ingresan en la hoja 'CARGA INICIAL' y se reflejan")
        print("automáticamente en la hoja 'Cobranzas' mediante fórmulas.")
        print("Para completar el cupón, debemos escribir en 'CARGA INICIAL'.")
    
    except Exception as e:
        print(f"Error al analizar el template: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analizar_template()

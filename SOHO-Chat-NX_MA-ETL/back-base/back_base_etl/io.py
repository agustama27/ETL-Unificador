"""Input/output functions for ETL source and target files."""

from __future__ import annotations

import os
import logging
import csv
import re
import unicodedata
from datetime import date
from collections.abc import Iterator
from typing import cast

import pandas as pd
from openpyxl import load_workbook

from .constants import (
    INPUT_COLUMN_ALIASES,
    INPUT_COLUMNS,
    INPUT_OPTIONAL_COLUMNS,
    OUTPUT_FILENAME_ROMAN,
    INPUT_SHEET_ALIASES,
    OUTPUT_FILENAME_PREFIX,
    PAGOS_REQUIRED_COLUMNS,
    PLANES_REQUIRED_COLUMNS,
)
from .validators import validate_required_columns


LOGGER = logging.getLogger("etl_naranjax")


ROMAN_PLAN_COLUMNS_TO_EXCLUDE = [f"plan_{idx}_cuotas" for idx in range(1, 8)]

ROMAN_OUTPUT_RENAME_MAP = {
    "plan_1_entrega": "monto_entrega_3",
    "plan_1_cuota_mensual": "monto_cuota_3",
    "plan_2_entrega": "monto_entrega_6",
    "plan_2_cuota_mensual": "monto_cuota_6",
    "plan_3_entrega": "monto_entrega_9",
    "plan_3_cuota_mensual": "monto_cuota_9",
    "plan_4_entrega": "monto_entrega_12",
    "plan_4_cuota_mensual": "monto_cuota_12",
    "plan_5_entrega": "monto_entrega_18",
    "plan_5_cuota_mensual": "monto_cuota_18",
    "plan_6_entrega": "monto_entrega_24",
    "plan_6_cuota_mensual": "monto_cuota_24",
    "plan_7_entrega": "monto_entrega_36",
    "plan_7_cuota_mensual": "monto_cuota_36",
}


def _format_roman_output_for_export(df: pd.DataFrame) -> pd.DataFrame:
    formatted = df.copy()

    formatted = formatted.drop(columns=ROMAN_PLAN_COLUMNS_TO_EXCLUDE, errors="ignore")
    formatted = formatted.rename(columns=ROMAN_OUTPUT_RENAME_MAP)

    if "tipo_marca_plan" in formatted.columns:
        mapped_tipo = formatted["tipo_marca_plan"].fillna("").astype(str).str.strip().str.lower()
        formatted["tipo_marca_plan"] = mapped_tipo.map({"con": "Con Plan", "sin": "Sin Plan"}).fillna("Sin Plan")

    return formatted


PLANES_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "nroproducto": ("nroproducto",),
    "nro_doc": ("nro_doc", "dni/nrodoc", "dni", "documento"),
    "cajon": ("cajon", "cajon_asignacion_cliente"),
    "deuda_total": ("deuda_total",),
    "deuda_vencida": ("deuda_vencida",),
    "plan": ("plan",),
    "importe_entrega": ("importe_entrega",),
    "importe_cuota": ("importe_cuota",),
}

PLANES_OPTIONAL_COLUMNS = {"importe_entrega", "nro_doc"}

PAGOS_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "nroproducto": ("nroproducto", "nro_producto", "producto", "documento", "dni"),
    "recupero": ("recupero",),
    "tipo_pago": ("tipo_pago", "producto", "proveedor"),
    "importe_pago": ("importe_pago", "importe", "importe_pagado"),
    "cajon_actual_prod": (
        "cajon_actual_prod",
        "cajon_act_prod",
        "cajon_asig_prod",
        "cajon_asignacion_cliente",
    ),
}

PAGOS_DEFAULT_VALUES: dict[str, str] = {
    "recupero": "",
    "tipo_pago": "",
    "importe_pago": "0",
    "cajon_actual_prod": "",
}


def _score_pagos_headers(headers: list[str]) -> int:
    positions_by_header: dict[str, str] = {}
    for header in headers:
        normalized_header = _normalize_column_name(header)
        if normalized_header and normalized_header not in positions_by_header:
            positions_by_header[normalized_header] = header

    score = 0
    for canonical in PAGOS_REQUIRED_COLUMNS:
        aliases = PAGOS_COLUMN_ALIASES.get(canonical, (canonical,))
        if any(_normalize_column_name(alias) in positions_by_header for alias in aliases):
            score += 1

    return score


def _detect_pagos_delimiter(filepath: str) -> str:
    with open(filepath, "r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        first_line = fh.readline()

    if first_line:
        best_delimiter = ";"
        best_score = -1
        for delimiter in (";", ","):
            parsed = next(csv.reader([first_line], delimiter=delimiter), [])
            score = _score_pagos_headers(parsed)
            if score > best_score or (score == best_score and len(parsed) > 1):
                best_score = score
                best_delimiter = delimiter

        if best_score > 0:
            return best_delimiter

    if sample:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
            return dialect.delimiter
        except csv.Error:
            pass

    return ";" if sample.count(";") >= sample.count(",") else ","


def _split_delimited_line_loose(line: str, delimiter: str) -> list[str]:
    fields: list[str] = []
    current: list[str] = []
    in_quotes = False
    idx = 0

    while idx < len(line):
        char = line[idx]

        if char == '"':
            if in_quotes and idx + 1 < len(line) and line[idx + 1] == '"':
                current.append('"')
                idx += 2
                continue
            in_quotes = not in_quotes
            idx += 1
            continue

        if char == delimiter and not in_quotes:
            fields.append("".join(current).strip())
            current = []
            idx += 1
            continue

        current.append(char)
        idx += 1

    fields.append("".join(current).strip())
    return [field.strip().strip('"').strip("'") for field in fields]


def _reparse_collapsed_pagos_csv(filepath: str, preferred_delimiter: str) -> tuple[pd.DataFrame, str] | None:
    with open(filepath, "r", encoding="utf-8-sig", errors="replace", newline="") as fh:
        raw_lines = [line.rstrip("\r\n") for line in fh if line.strip()]

    if not raw_lines:
        return None

    header_line = raw_lines[0]
    best_headers: list[str] | None = None
    best_delimiter: str | None = None
    best_score = -1

    candidates: list[str] = []
    for delimiter in (preferred_delimiter, ",", ";"):
        if delimiter not in candidates:
            candidates.append(delimiter)

    header_variants = [header_line]
    if header_line.startswith('"') and header_line.endswith('"') and len(header_line) > 1:
        header_variants.append(header_line[1:-1])

    for delimiter in candidates:
        for header_variant in header_variants:
            headers = _split_delimited_line_loose(header_variant, delimiter)
            if len(headers) <= 1:
                continue
            score = _score_pagos_headers(headers)
            if score > best_score or (score == best_score and best_headers is not None and len(headers) > len(best_headers)):
                best_headers = headers
                best_delimiter = delimiter
                best_score = score

    if best_headers is None or best_delimiter is None:
        return None

    records: list[list[str]] = []
    expected_columns = len(best_headers)
    for line in raw_lines[1:]:
        values = _split_delimited_line_loose(line, best_delimiter)
        if len(values) < expected_columns:
            values.extend([""] * (expected_columns - len(values)))
        elif len(values) > expected_columns:
            values = values[:expected_columns]
        records.append(values)

    reparsed = pd.DataFrame.from_records(records, columns=[str(column) for column in best_headers])
    return reparsed, best_delimiter


def _normalize_column_name(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    normalized = unicodedata.normalize("NFKD", text)
    no_accents = "".join(char for char in normalized if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]", "", no_accents)


def _normalize_sheet_name(value: str) -> str:
    return value.strip().lower().replace("ó", "o")


def _resolve_input_sheet_name(filepath: str) -> str:
    workbook = load_workbook(filename=filepath, read_only=True, data_only=True)
    try:
        available_sheets = workbook.sheetnames
    finally:
        workbook.close()

    for alias in INPUT_SHEET_ALIASES:
        if alias in available_sheets:
            return alias

    pattern = re.compile(r"^asignacion\s+m90\s+-\s+.+")
    for sheet_name in available_sheets:
        normalized = _normalize_sheet_name(sheet_name)
        if pattern.match(normalized):
            return sheet_name

    available = ", ".join(available_sheets) or "(none)"
    aliases = ", ".join(INPUT_SHEET_ALIASES)
    raise ValueError(
        "Input workbook does not contain a valid base mensual sheet. "
        f"Expected one of: {aliases}, or pattern 'Asignacion M90 - *'. "
        f"Available sheets: {available}"
    )


def _resolve_planes_column_mapping(headers: list[str], context: str) -> dict[str, int]:
    positions_by_header: dict[str, int] = {}
    for idx, header in enumerate(headers):
        if header and header not in positions_by_header:
            positions_by_header[header] = idx

    missing_required: list[str] = []
    mapping: dict[str, int] = {}

    for canonical, aliases in PLANES_COLUMN_ALIASES.items():
        found_idx = next(
            (
                positions_by_header[normalized_alias]
                for alias in aliases
                if (normalized_alias := _normalize_column_name(alias)) in positions_by_header
            ),
            None,
        )
        if found_idx is None:
            if canonical in PLANES_OPTIONAL_COLUMNS:
                continue
            missing_required.append(canonical)
            continue
        mapping[canonical] = found_idx

    if "nroproducto" not in mapping and "nro_doc" in mapping:
        mapping["nroproducto"] = mapping["nro_doc"]
        missing_required = [column for column in missing_required if column != "nroproducto"]

    if missing_required:
        available = ", ".join(headers) or "(none)"
        raise ValueError(
            "Unsupported PLANES format: could not map required columns "
            f"{', '.join(sorted(missing_required))}. Available columns: {available}"
        )

    return mapping


def _normalize_planes_df(df: pd.DataFrame, context: str) -> pd.DataFrame:
    headers = [_normalize_column_name(column) for column in df.columns]
    mapping = _resolve_planes_column_mapping(headers, context=context)

    result = pd.DataFrame()
    for canonical in PLANES_REQUIRED_COLUMNS:
        if canonical in mapping:
            result[canonical] = df.iloc[:, mapping[canonical]]
        elif canonical in PLANES_OPTIONAL_COLUMNS:
            result[canonical] = 0

    validate_required_columns(result, PLANES_REQUIRED_COLUMNS, context=context)
    if "nro_doc" in mapping:
        result["nro_doc"] = df.iloc[:, mapping["nro_doc"]]
    else:
        result["nro_doc"] = ""
    return result


def load_input(filepath: str) -> pd.DataFrame:
    """Load input Excel from base mensual sheet using header aliases."""
    sheet_name = _resolve_input_sheet_name(filepath)
    df = pd.read_excel(filepath, sheet_name=sheet_name, engine="openpyxl", header=0)

    positions_by_header: dict[str, int] = {}
    original_headers: list[str] = []
    for idx, header in enumerate(df.columns):
        original = str(header).strip() if header is not None else ""
        original_headers.append(original)
        normalized = _normalize_column_name(header)
        if normalized and normalized not in positions_by_header:
            positions_by_header[normalized] = idx

    mapping: dict[str, int] = {}
    missing_required: list[str] = []
    for canonical in INPUT_COLUMNS:
        aliases = INPUT_COLUMN_ALIASES.get(canonical, (canonical,))
        found_idx = next((positions_by_header[_normalize_column_name(alias)] for alias in aliases if _normalize_column_name(alias) in positions_by_header), None)
        if found_idx is None:
            if canonical in INPUT_OPTIONAL_COLUMNS:
                continue
            missing_required.append(canonical)
            continue
        mapping[canonical] = found_idx

    if missing_required:
        available = ", ".join(header for header in original_headers if header) or "(none)"
        details = "; ".join(
            f"{canonical} (aliases: {', '.join(INPUT_COLUMN_ALIASES.get(canonical, (canonical,)))})"
            for canonical in sorted(missing_required)
        )
        raise ValueError(
            "Unsupported base mensual format: could not map required columns "
            f"{', '.join(sorted(missing_required))}. Missing details: {details}. "
            f"Available columns: {available}"
        )

    normalized = pd.DataFrame(index=df.index)
    for canonical in INPUT_COLUMNS:
        if canonical in mapping:
            normalized[canonical] = df.iloc[:, mapping[canonical]]
        elif canonical in INPUT_OPTIONAL_COLUMNS:
            normalized[canonical] = pd.NA
    return normalized


def load_planes(filepath: str) -> pd.DataFrame:
    """Load daily plans file and validate required columns."""
    df = pd.read_excel(filepath, sheet_name="default_1", engine="openpyxl")
    normalized = _normalize_planes_df(df, context="PLANES file")
    LOGGER.info("Loaded PLANES file rows=%s path=%s", len(normalized), filepath)
    return normalized


def iter_planes_chunks(filepath: str, chunk_size: int = 50_000) -> Iterator[pd.DataFrame]:
    """Yield PLANES rows in chunks using openpyxl read-only streaming.

    This avoids loading the full workbook into a single pandas dataframe and
    keeps memory usage stable for large daily files.
    """
    if chunk_size <= 0:
        raise ValueError("chunk_size must be greater than zero")

    workbook = load_workbook(filename=filepath, read_only=True, data_only=True)
    try:
        if "default_1" not in workbook.sheetnames:
            raise ValueError("PLANES file must contain sheet 'default_1'")

        sheet = workbook["default_1"]
        rows = sheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if raw_headers is None:
            empty = pd.DataFrame(columns=pd.Index(PLANES_REQUIRED_COLUMNS))
            validate_required_columns(empty, PLANES_REQUIRED_COLUMNS, context="PLANES file")
            return

        headers = [_normalize_column_name(column) for column in raw_headers]
        positions = _resolve_planes_column_mapping(headers, context="PLANES file")
        buffer: list[dict[str, object]] = []

        for row in rows:
            if row is None:
                continue

            record: dict[str, object] = {}
            for column in [*PLANES_REQUIRED_COLUMNS, "nro_doc"]:
                idx = positions.get(column)
                if idx is None:
                    record[column] = 0 if column in PLANES_OPTIONAL_COLUMNS else None
                    continue
                value = row[idx] if idx < len(row) else None
                record[column] = cast(object, value)
            buffer.append(record)

            if len(buffer) >= chunk_size:
                yield pd.DataFrame.from_records(buffer, columns=pd.Index([*PLANES_REQUIRED_COLUMNS, "nro_doc"]))
                buffer = []

        if buffer:
            yield pd.DataFrame.from_records(buffer, columns=pd.Index([*PLANES_REQUIRED_COLUMNS, "nro_doc"]))
    finally:
        workbook.close()


def load_pagos(filepath: str) -> pd.DataFrame:
    """Load daily payments file supporting legacy and evoltis formats."""
    detected_delimiter = _detect_pagos_delimiter(filepath)

    df = pd.read_csv(
        filepath,
        sep=detected_delimiter,
        dtype=str,
        keep_default_na=False,
        encoding="utf-8-sig",
        skipinitialspace=True,
    )

    should_try_reparse = False
    if len(df.columns) == 1:
        collapsed_header = str(df.columns[0])
        should_try_reparse = any(delimiter in collapsed_header for delimiter in (",", ";"))
    elif df.empty and len(df.columns) > 1:
        suspicious_header = any(('"' in str(column)) or ("\r" in str(column)) or ("\n" in str(column)) for column in df.columns)
        should_try_reparse = suspicious_header

    if should_try_reparse:
        reparsed = _reparse_collapsed_pagos_csv(filepath, preferred_delimiter=detected_delimiter)
        if reparsed is not None:
            df, detected_delimiter = reparsed

    df.columns = [str(column).strip().strip('"').strip("'").lower() for column in df.columns]

    positions_by_header: dict[str, str] = {}
    for header in df.columns:
        normalized_header = _normalize_column_name(header)
        if normalized_header and normalized_header not in positions_by_header:
            positions_by_header[normalized_header] = header

    normalized = pd.DataFrame(index=df.index)
    missing_required: list[str] = []
    for canonical in PAGOS_REQUIRED_COLUMNS:
        aliases = PAGOS_COLUMN_ALIASES.get(canonical, (canonical,))
        found_header = next(
            (
                positions_by_header[normalized_alias]
                for alias in aliases
                if (normalized_alias := _normalize_column_name(alias)) in positions_by_header
            ),
            None,
        )
        if found_header is not None:
            normalized[canonical] = df[found_header].astype(str).str.strip()
            continue

        if canonical in PAGOS_DEFAULT_VALUES:
            normalized[canonical] = PAGOS_DEFAULT_VALUES[canonical]
            continue

        missing_required.append(canonical)

    if missing_required:
        available = ", ".join(str(column) for column in df.columns) or "(none)"
        raise ValueError(
            "Unsupported PAGOS format: could not map required columns "
            f"{', '.join(sorted(missing_required))}. Available columns: {available}"
        )

    if "importe_pago" not in normalized.columns:
        normalized["importe_pago"] = "0"

    for raw_column in ("importe_pago", "importe_pago_1"):
        if raw_column in df.columns:
            normalized["importe_pago"] = df[raw_column].astype(str).str.strip()
            break
    else:
        for raw_column in ("importe_pago", "importe", "importe_pagado"):
            if raw_column in df.columns:
                normalized["importe_pago"] = df[raw_column].astype(str).str.strip()
                break

    validate_required_columns(normalized, PAGOS_REQUIRED_COLUMNS, context="PAGOS file")
    LOGGER.info(
        "Loaded PAGOS file rows=%s path=%s delimiter=%s",
        len(normalized),
        filepath,
        detected_delimiter,
    )
    return normalized


def save_output(
    df: pd.DataFrame,
    output_dir: str,
    prefix: str = OUTPUT_FILENAME_PREFIX,
    date_format: str = "%Y%m%d",
    suffix: str = ".csv",
) -> str:
    """Save transformed dataframe to CSV with expected naming/format."""
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{prefix}{date.today().strftime(date_format)}{suffix}"
    output_path = os.path.join(output_dir, filename)

    export_df = _format_roman_output_for_export(df) if prefix == OUTPUT_FILENAME_ROMAN else df

    export_df.to_csv(
        output_path,
        sep=";",
        encoding="utf-8",
        index=False,
        lineterminator="\n",
    )
    return output_path
